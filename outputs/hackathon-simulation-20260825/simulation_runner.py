from __future__ import annotations

import json
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import httpx
from openpyxl import load_workbook


BASE = "http://127.0.0.1:8010"
OUT = Path(__file__).resolve().parent
STATE_PATH = OUT / "simulation_state.json"
REGISTRATION = OUT / "simulation_registration_30_teams.xlsx"
ROUND1 = OUT / "simulation_round1_problems.xlsx"
WILDCARD = OUT / "simulation_wildcard_problems.xlsx"
CREDENTIALS = OUT / "simulation_registration_credentials.xlsx"
FINAL_EXPORT = OUT / "simulation_final_registration_export.xlsx"


def fail(message: str, response: httpx.Response | None = None):
    detail = f"{message}"
    if response is not None:
        detail += f" | HTTP {response.status_code}: {response.text[:1000]}"
    raise RuntimeError(detail)


class Api:
    def __init__(self, token: str | None = None):
        self.client = httpx.Client(base_url=BASE, timeout=30)
        if token:
            self.client.headers["Authorization"] = f"Bearer {token}"

    def request(self, method: str, path: str, *, expected=(200,), **kwargs):
        response = self.client.request(method, path, **kwargs)
        if response.status_code not in expected:
            fail(f"{method} {path} failed", response)
        return response

    def get(self, path: str, **kwargs):
        return self.request("GET", path, **kwargs)

    def post(self, path: str, **kwargs):
        return self.request("POST", path, **kwargs)

    def put(self, path: str, **kwargs):
        return self.request("PUT", path, **kwargs)


def login(email: str, password: str, *, leaderboard=False) -> str:
    path = "/leaderboard/login" if leaderboard else "/login"
    response = httpx.post(f"{BASE}{path}", data={"username": email, "password": password}, timeout=30)
    if response.status_code != 200:
        fail(f"Login failed for {email}", response)
    return response.json()["access_token"]


def load_state() -> dict:
    return json.loads(STATE_PATH.read_text(encoding="utf-8"))


def save_state(state: dict):
    STATE_PATH.write_text(json.dumps(state, indent=2), encoding="utf-8")


def workbook_credentials() -> list[dict]:
    workbook = load_workbook(CREDENTIALS, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    normalized = {"".join(ch for ch in header.lower() if ch.isalnum()): index for index, header in enumerate(headers)}
    email_index = next(index for key, index in normalized.items() if key in {"leaderemail", "leaderloginemail", "email"})
    password_index = next(index for key, index in normalized.items() if "password" in key)
    team_index = next(index for key, index in normalized.items() if key == "teamname")
    credentials = []
    for row in rows[1:]:
        if not row[team_index]:
            continue
        credentials.append({
            "team": str(row[team_index]),
            "email": str(row[email_index]),
            "password": str(row[password_index]),
        })
    workbook.close()
    return credentials


def phase_setup():
    health = httpx.get(f"{BASE}/health", timeout=30)
    if health.status_code != 200:
        fail("Backend health failed", health)
    admin_token = login("admin.demo@bidtobuild.example.com", "DemoAdmin@123")
    display_token = login("leaderboard@bidtobuild.example.com", "Leaderboard@123", leaderboard=True)
    admin = Api(admin_token)

    config = admin.put("/admin/config", json={
        "round1_preview_seconds": 300,
        "round1_bid_seconds": 300,
        "round1_winner_count": 5,
        "round1_minimum_bid": 25,
        "wildcard_enabled": True,
        "wildcard_slots": 3,
        "wildcard_application_seconds": 300,
        "wildcard_bid_seconds": 300,
        "wildcard_selection_seconds": 5,
        "wildcard_starting_bid": 150,
        "coding_duration_seconds": 300,
        "bid_cooldown_seconds": 5,
    }).json()

    with REGISTRATION.open("rb") as handle:
        imported = admin.post(
            "/admin/registration/import",
            files={"file": (REGISTRATION.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).json()
    download = admin.get(f"/admin/registration/import/download/{imported['download_token']}")
    CREDENTIALS.write_bytes(download.content)
    credentials = workbook_credentials()
    if len(credentials) != 30:
        fail(f"Expected 30 generated credentials, got {len(credentials)}")

    leaders = []
    for credential in credentials:
        token = login(credential["email"], credential["password"])
        dashboard = Api(token).get("/participant/dashboard").json()
        leaders.append({
            **credential,
            "token": token,
            "team_id": dashboard["team"]["id"],
        })

    with ROUND1.open("rb") as handle:
        admin.post(
            "/admin/rounds/round-1/problems/import",
            files={"file": (ROUND1.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    round_payload = admin.get("/admin/rounds/round-1").json()
    problems = sorted(round_payload["problems"], key=lambda item: int(str(item.get("problem_number") or item.get("number")).split("-")[-1]))
    if len(problems) != 6:
        fail(f"Expected 6 Round 1 problems, got {len(problems)}")

    first = problems[0]
    admin.post(f"/admin/rounds/round-1/problems/{first['id']}/select")
    admin.post("/admin/rounds/round-1/preview/start")
    admin.post("/admin/rounds/round-1/bidding/start")
    increments = [5, 10, 25, 5, 10]
    bid_results = []
    for leader, increment in zip(leaders[:5], increments):
        bid_results.append(Api(leader["token"]).post("/bid", json={"ps_id": first["id"], "increment": increment}).json())

    cooldown = Api(leaders[0]["token"]).post(
        "/bid", json={"ps_id": first["id"], "increment": 5}, expected=(429,)
    )
    arbitrary = Api(leaders[5]["token"]).post(
        "/bid", json={"ps_id": first["id"], "amount": 1}, expected=(422,)
    )
    board = Api(display_token).get("/leaderboard/round-1").json()
    rejected_absent = all(row.get("team_name") != leaders[5]["team"] for row in board["rows"])

    state = {
        "admin_token": admin_token,
        "display_token": display_token,
        "config": config,
        "import": imported,
        "leaders": leaders,
        "round1_problems": problems,
        "checks": {
            "admin_login": True,
            "display_login": True,
            "leader_logins": 30,
            "round1_cooldown_status": cooldown.status_code,
            "arbitrary_bid_status": arbitrary.status_code,
            "rejected_bid_absent": rejected_absent,
            "first_live_board_rows": len(board["rows"]),
        },
    }
    save_state(state)
    print(json.dumps({"phase": "setup", "live_round1_problem": first, "board_rows": len(board["rows"]), "credentials": len(credentials)}, indent=2))


def close_assign(admin: Api):
    admin.post("/admin/rounds/round-1/bidding/close")
    return admin.post("/admin/rounds/round-1/assign-winners").json()


def phase_finish():
    state = load_state()
    admin = Api(state["admin_token"])
    leaders = state["leaders"]
    problems = state["round1_problems"]
    round_winners = []

    result = close_assign(admin)
    round_winners.append(result["winners"])
    assigned_retry = Api(leaders[0]["token"]).post(
        "/bid", json={"ps_id": problems[0]["id"], "increment": 5}, expected=(409,)
    )

    increments = [5, 10, 25, 5, 10]
    for problem_index in range(1, 6):
        problem = problems[problem_index]
        admin.post(f"/admin/rounds/round-1/problems/{problem['id']}/select")
        admin.post("/admin/rounds/round-1/preview/start")
        admin.post("/admin/rounds/round-1/bidding/start")
        start = problem_index * 5
        for leader, increment in zip(leaders[start:start + 5], increments):
            Api(leader["token"]).post("/bid", json={"ps_id": problem["id"], "increment": increment})
        result = close_assign(admin)
        round_winners.append(result["winners"])

    dashboards = [Api(leader["token"]).get("/participant/dashboard").json() for leader in leaders]
    round1_assigned = sum(bool(row.get("round1Problem")) for row in dashboards)
    unique_round1 = len({row["round1Problem"]["id"] for row in dashboards if row.get("round1Problem")})
    admin.post("/admin/rounds/round-1/end")
    admin.post("/admin/rounds/wildcard/applications/open")

    applicants = leaders[:8]
    for leader in applicants:
        Api(leader["token"]).post("/wildcard/apply")
    duplicate_application = Api(applicants[0]["token"]).post("/wildcard/apply").json()

    # Re-importing the same registration is idempotent for team/account identity.
    with REGISTRATION.open("rb") as handle:
        duplicate_import = admin.post(
            "/admin/registration/import",
            files={"file": (REGISTRATION.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).json()
    current_team_ids = [Api(leader["token"]).get("/participant/dashboard").json()["team"]["id"] for leader in leaders]
    duplicate_import_stable = current_team_ids == [leader["team_id"] for leader in leaders]

    admin.post("/admin/rounds/wildcard/applications/close")
    late_application = Api(leaders[8]["token"]).post("/wildcard/apply", expected=(409,))

    with WILDCARD.open("rb") as handle:
        admin.post(
            "/admin/rounds/wildcard/problems/import",
            files={"file": (WILDCARD.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    slots_too_many = admin.post("/admin/rounds/wildcard/slots", json={"slots": 4}, expected=(422,))
    admin.post("/admin/rounds/wildcard/slots", json={"slots": 3})

    # Prove server-side balance enforcement, then restore the live starting price before real bids.
    admin.put("/admin/config", json={"wildcard_starting_bid": 990})
    admin.post("/admin/rounds/wildcard/bidding/start")
    non_applicant = Api(leaders[8]["token"]).post("/wildcard/bid", json={"increment": 5}, expected=(403,))
    over_balance = Api(applicants[7]["token"]).post("/wildcard/bid", json={"increment": 25}, expected=(400,))
    admin.put("/admin/config", json={"wildcard_starting_bid": 150})

    bid_order = list(reversed(applicants))
    wildcard_increments = [5, 10, 25, 5, 10, 25, 5, 10]
    for leader, increment in zip(bid_order, wildcard_increments):
        Api(leader["token"]).post("/wildcard/bid", json={"increment": increment})
    cooldown_two_sessions = Api(applicants[0]["token"]).post("/wildcard/bid", json={"increment": 5}, expected=(429,))
    wildcard_board = Api(state["display_token"]).get("/leaderboard/wildcard").json()
    closed = admin.post("/admin/rounds/wildcard/bidding/close").json()
    winners = closed["winners"]
    if len(winners) != 3:
        fail(f"Expected 3 Wildcard winners, got {len(winners)}")
    ranked_ids = [winner["team_id"] for winner in winners]
    ranked = [next(leader for leader in leaders if leader["team_id"] == team_id) for team_id in ranked_ids]

    # Rank 2 cannot select while Rank 1 is active.
    rank1_choices = Api(ranked[0]["token"]).get("/participant/problems?round=2").json()
    rank2_choices_waiting = Api(ranked[1]["token"]).get("/participant/problems?round=2").json()
    out_of_turn = Api(ranked[1]["token"]).post(f"/wildcard/select/{rank1_choices[0]['id']}", expected=(409,))

    # Two sessions submit the same choice concurrently; exactly one assignment must win.
    target_problem = rank1_choices[-1]["id"]
    def concurrent_select():
        response = httpx.post(
            f"{BASE}/wildcard/select/{target_problem}",
            headers={"Authorization": f"Bearer {ranked[0]['token']}"}, timeout=30,
        )
        return response.status_code
    with ThreadPoolExecutor(max_workers=2) as pool:
        concurrent_statuses = sorted(list(pool.map(lambda _: concurrent_select(), range(2))))

    rank2_choices = Api(ranked[1]["token"]).get("/participant/problems?round=2").json()
    # No browser/request is required for timeout: the existing background worker advances the turn.
    time.sleep(6.5)
    rank2_status = Api(ranked[1]["token"]).get("/wildcard/status").json()
    rank3_status = Api(ranked[2]["token"]).get("/wildcard/status").json()
    rank3_choices = Api(ranked[2]["token"]).get("/participant/problems?round=2").json()
    if len(rank3_choices) != 1:
        fail(f"Rank 3 expected one remaining problem, got {len(rank3_choices)}")
    rank3_manual = Api(ranked[2]["token"]).post(f"/wildcard/select/{rank3_choices[0]['id']}").json()

    wildcard_dashboards = [Api(leader["token"]).get("/participant/dashboard").json() for leader in ranked]
    wildcard_problem_ids = [row["wildcardProblem"]["id"] for row in wildcard_dashboards]
    round1_history_preserved = all(row.get("round1Problem") for row in wildcard_dashboards)

    admin.post("/admin/submissions/open")
    for index, leader in enumerate(leaders, start=1):
        Api(leader["token"]).put("/submissions/me", json={"repository_url": f"https://github.com/bidtobuild-simulation/team-{index:02d}"})
    updated_url = "https://github.com/bidtobuild-simulation/team-01-updated"
    updated_submission = Api(leaders[0]["token"]).put("/submissions/me", json={"repository_url": updated_url}).json()
    submissions_before_close = admin.get("/admin/submissions").json()
    imported_submission_rows = [row for row in submissions_before_close["rows"] if row["team_name"].startswith("Team ")]
    admin.post("/admin/submissions/close")
    late_submission = Api(leaders[1]["token"]).put(
        "/submissions/me", json={"repository_url": "https://github.com/bidtobuild-simulation/late"}, expected=(409,)
    )

    judging = admin.get("/admin/judging").json()
    winner_payload = {
        "first_place_team_id": leaders[0]["team_id"],
        "second_place_team_id": leaders[1]["team_id"],
        "third_place_team_id": leaders[2]["team_id"],
    }
    saved = admin.put("/admin/judging/winners", json=winner_payload).json()
    unpublished = Api(state["display_token"]).get("/public/leaderboard").json()

    export = admin.get("/admin/registration/assignments")
    FINAL_EXPORT.write_bytes(export.content)

    display_forbidden = Api(state["display_token"]).get("/admin/config", expected=(403,))
    participant_forbidden = Api(leaders[0]["token"]).get("/admin/config", expected=(403,))
    logout_token = login(leaders[29]["email"], leaders[29]["password"])
    logout_api = Api(logout_token)
    logout_api.post("/logout")
    old_token = logout_api.get("/participant/dashboard", expected=(401,))

    state["round1_winners"] = round_winners
    state["wildcard_winners"] = winners
    state["winner_payload"] = winner_payload
    state["checks"].update({
        "assigned_team_rebid_status": assigned_retry.status_code,
        "round1_assigned": round1_assigned,
        "round1_problem_count": unique_round1,
        "duplicate_application": duplicate_application.get("message"),
        "duplicate_import_teams_created": duplicate_import.get("teams_created"),
        "duplicate_import_stable": duplicate_import_stable,
        "late_application_status": late_application.status_code,
        "slots_too_many_status": slots_too_many.status_code,
        "non_applicant_bid_status": non_applicant.status_code,
        "over_balance_status": over_balance.status_code,
        "wildcard_cooldown_status": cooldown_two_sessions.status_code,
        "wildcard_board_rows": len(wildcard_board["rows"]),
        "wildcard_pool_rank1": len(rank1_choices),
        "wildcard_pool_rank2": len(rank2_choices),
        "wildcard_pool_rank3": len(rank3_choices),
        "rank2_waiting_choices": len(rank2_choices_waiting),
        "out_of_turn_status": out_of_turn.status_code,
        "concurrent_select_statuses": concurrent_statuses,
        "rank2_selection_method": rank2_status.get("selection_method"),
        "rank3_fresh_turn": rank3_status.get("is_selection_turn"),
        "rank3_manual_method": rank3_manual.get("selection_method"),
        "wildcard_unique_assignments": len(set(wildcard_problem_ids)),
        "round1_history_preserved": round1_history_preserved,
        "submissions": len(imported_submission_rows),
        "updated_submission": updated_submission.get("repository_url") == updated_url,
        "late_submission_status": late_submission.status_code,
        "judging_team_count": len([row for row in judging["teams"] if row["team_name"].startswith("Team ")]),
        "saved_unpublished": saved.get("result_status") == "WAITING" and unpublished.get("results") is None,
        "display_forbidden_status": display_forbidden.status_code,
        "participant_forbidden_status": participant_forbidden.status_code,
        "old_token_status": old_token.status_code,
    })
    save_state(state)
    print(json.dumps({
        "phase": "finish",
        "round1_assigned": round1_assigned,
        "wildcard_winners": [winner["team_name"] for winner in winners],
        "selection_counts": [len(rank1_choices), len(rank2_choices), len(rank3_choices)],
        "submissions": len(imported_submission_rows),
        "results_status": saved["result_status"],
    }, indent=2))


def phase_publish():
    state = load_state()
    admin = Api(state["admin_token"])
    published = admin.post("/admin/judging/publish").json()
    public = Api(state["display_token"]).get("/public/leaderboard").json()
    participant = Api(state["leaders"][0]["token"]).get("/participant/dashboard").json()
    state["checks"].update({
        "published_status": published.get("result_status"),
        "public_mode_after_publish": public.get("mode"),
        "participant_results_visible": participant.get("finalResults") is not None,
    })
    save_state(state)
    print(json.dumps({"phase": "publish", "public_mode": public.get("mode"), "participant_results": participant.get("finalResults")}, indent=2))


if __name__ == "__main__":
    phase = sys.argv[1] if len(sys.argv) > 1 else "setup"
    {"setup": phase_setup, "finish": phase_finish, "publish": phase_publish}[phase]()
