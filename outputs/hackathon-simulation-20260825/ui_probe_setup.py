from io import BytesIO
import json
from pathlib import Path

import httpx
from openpyxl import load_workbook

BASE = "http://127.0.0.1:8011"
OUT = Path(__file__).resolve().parent

def request(client, method, path, **kwargs):
    response = client.request(method, f"{BASE}{path}", timeout=30, **kwargs)
    if response.status_code != 200:
        raise RuntimeError(f"{method} {path}: {response.status_code} {response.text[:500]}")
    return response

def login(email, password, leaderboard=False):
    path = "/leaderboard/login" if leaderboard else "/login"
    response = httpx.post(f"{BASE}{path}", data={"username": email, "password": password}, timeout=30)
    if response.status_code != 200:
        raise RuntimeError(f"login {email}: {response.status_code}")
    return response.json()["access_token"]

admin_token = login("admin.demo@bidtobuild.example.com", "DemoAdmin@123")
admin = httpx.Client(headers={"Authorization": f"Bearer {admin_token}"})
request(admin, "PUT", "/admin/config", json={"round1_preview_seconds": 300, "round1_bid_seconds": 300, "bid_cooldown_seconds": 5})
with (OUT / "simulation_registration_30_teams.xlsx").open("rb") as handle:
    imported = request(admin, "POST", "/admin/registration/import", files={"file": ("simulation_registration_30_teams.xlsx", handle)}).json()
credential_bytes = request(admin, "GET", f"/admin/registration/import/download/{imported['download_token']}").content
workbook = load_workbook(BytesIO(credential_bytes), read_only=True, data_only=True)
rows = list(workbook.active.iter_rows(values_only=True))
headers = [str(value or "") for value in rows[0]]
email_index = headers.index("Leader Login Email")
password_index = headers.index("Leader Password")
leaders = [{"email": str(row[email_index]), "password": str(row[password_index])} for row in rows[1:] if row[email_index]]
workbook.close()
with (OUT / "simulation_round1_problems.xlsx").open("rb") as handle:
    request(admin, "POST", "/admin/rounds/round-1/problems/import", files={"file": ("simulation_round1_problems.xlsx", handle)})
round_data = request(admin, "GET", "/admin/rounds/round-1").json()
problem = sorted(round_data["problems"], key=lambda row: int(row["problem_number"]))[0]
request(admin, "POST", f"/admin/rounds/round-1/problems/{problem['id']}/select")
request(admin, "POST", "/admin/rounds/round-1/preview/start")
request(admin, "POST", "/admin/rounds/round-1/bidding/start")
for leader, increment in zip(leaders[:5], [5, 10, 25, 5, 10]):
    token = login(leader["email"], leader["password"])
    request(httpx.Client(headers={"Authorization": f"Bearer {token}"}), "POST", "/bid", json={"ps_id": problem["id"], "increment": increment})
(OUT / "ui_probe_login.json").write_text(json.dumps({"leader": leaders[0], "problem": problem}), encoding="utf-8")
print(json.dumps({"teams": len(leaders), "problem": problem, "bids": 5}))
