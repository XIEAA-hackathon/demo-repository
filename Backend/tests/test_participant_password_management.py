import csv
import io
import json

from app.models.models import GameConfig, RegistrationImportRow, User


def _registration_csv(
    leader_password: str,
    member_password: str = "MemberInitial@123",
    *,
    team_name: str = "Password Team",
    leader_email: str = "leader.test@example.com",
) -> bytes:
    return (
        "Team Name,Leader Name,Leader Email,Leader Password,"
        "Member 1 Name,Member 1 Email,Member 1 Password\n"
        f"{team_name},Test Leader,{leader_email},{leader_password},"
        f"Test Member,member.test@example.com,{member_password}\n"
    ).encode()


def _login(client, login_id: str, password: str):
    return client.post("/login", data={"username": login_id, "password": password})


def _import(client, admin_headers, content: bytes):
    return client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("participants.csv", content, "text/csv")},
    )


def _set_password(client, admin_headers, user_id: int, password: str):
    return client.put(
        f"/admin/registration/participant-accounts/{user_id}/password",
        headers=admin_headers,
        json={"new_password": password, "confirm_password": password},
    )


def test_imported_password_lifecycle_and_resets(client, admin_headers, db):
    initial = _import(client, admin_headers, _registration_csv("Alpha@123"))
    assert initial.status_code == 200, initial.text
    assert initial.json()["leaders_created"] == 1
    assert initial.json()["participant_accounts_created"] == 1

    immediate = client.get(
        f"/admin/registration/import/download/{initial.json()['download_token']}",
        headers=admin_headers,
    )
    row = next(csv.DictReader(io.StringIO(immediate.content.decode("utf-8-sig"))))
    assert row["Leader Password"] == "Alpha@123"
    assert row["Member 1 Password"] == "MemberInitial@123"
    assert row["Credential Status"] == "PASSWORD SET"
    initial_login = _login(client, "leader.test@example.com", "Alpha@123")
    assert initial_login.status_code == 200
    initial_headers = {"Authorization": f"Bearer {initial_login.json()['access_token']}"}
    assert _login(client, "member.test@example.com", "MemberInitial@123").status_code == 200

    stored = db.query(RegistrationImportRow).one()
    assert "Alpha@123" not in stored.source_values_json
    assert "MemberInitial@123" not in stored.source_values_json
    assert "password" not in json.dumps(json.loads(stored.members_json)).lower()

    changed = _import(client, admin_headers, _registration_csv("AlphaNew@456"))
    assert changed.status_code == 200, changed.text
    assert client.get("/participant/dashboard", headers=initial_headers).status_code == 401
    assert _login(client, "leader.test@example.com", "Alpha@123").status_code == 401
    assert _login(client, "leader.test@example.com", "AlphaNew@456").status_code == 200

    blank = _import(client, admin_headers, _registration_csv("", ""))
    assert blank.status_code == 200, blank.text
    assert blank.json()["rows_failed"] == 0
    assert _login(client, "leader.test@example.com", "AlphaNew@456").status_code == 200
    assert _login(client, "member.test@example.com", "MemberInitial@123").status_code == 200

    accounts = client.get(
        "/admin/registration/participant-accounts", headers=admin_headers,
    ).json()["accounts"]
    leader = next(account for account in accounts if account["login_id"] == "leader.test@example.com")
    manual = _set_password(client, admin_headers, leader["user_id"], "Manual@789")
    assert manual.status_code == 200, manual.text
    assert _login(client, "leader.test@example.com", "AlphaNew@456").status_code == 401
    assert _login(client, "leader.test@example.com", "Manual@789").status_code == 200

    reset_event = client.post(
        "/admin/event-data/reset",
        headers=admin_headers,
        json={"confirmation": "RESET EVENT"},
    )
    assert reset_event.status_code == 200, reset_event.text
    assert _login(client, "leader.test@example.com", "AlphaNew@456").status_code == 401
    assert _login(client, "leader.test@example.com", "Manual@789").status_code == 200

    db.query(GameConfig).one().state = "SUBMISSION"
    db.commit()
    reset_credentials = client.post(
        "/admin/registration/credentials/reset",
        headers=admin_headers,
        json={"confirmation": "RESET CREDENTIALS"},
    )
    assert reset_credentials.status_code == 200, reset_credentials.text
    assert reset_credentials.json()["event_state"] == "SUBMISSION"
    assert _login(client, "leader.test@example.com", "Manual@789").status_code == 401

    restored = _import(client, admin_headers, _registration_csv("Restored@123"))
    assert restored.status_code == 200, restored.text
    assert _login(client, "leader.test@example.com", "Restored@123").status_code == 200
    db.expire_all()
    assert db.query(GameConfig).one().state == "SUBMISSION"


def test_new_account_with_blank_password_is_rejected_per_row(client, admin_headers, db):
    response = _import(
        client,
        admin_headers,
        _registration_csv(
            "",
            "",
            team_name="Blank Password Team",
            leader_email="blank.password@example.com",
        ),
    )
    assert response.status_code == 200, response.text
    summary = response.json()
    assert summary["teams_processed"] == 0
    assert summary["rows_failed"] == 1
    messages = [error["message"] for error in summary["errors"]]
    assert "Leader Password is required for a new participant account." in messages
    assert "Member 1 Password is required for a new participant account." in messages
    assert db.query(User).filter(User.email == "blank.password@example.com").count() == 0


def test_later_assignment_export_does_not_expose_imported_passwords(client, admin_headers):
    imported = _import(client, admin_headers, _registration_csv("Alpha@123"))
    assert imported.status_code == 200, imported.text
    export = client.get("/admin/registration/assignments", headers=admin_headers)
    assert export.status_code == 200, export.text
    row = next(csv.DictReader(io.StringIO(export.content.decode("utf-8-sig"))))
    assert row["Leader Password"] == "NOT EXPORTED"
    assert row["Member 1 Password"] == "NOT EXPORTED"


def test_sample_and_xlsx_import_include_and_preserve_password(client, admin_headers):
    sample = client.get("/admin/registration/sample.csv", headers=admin_headers)
    assert sample.status_code == 200
    assert "Leader Password" in next(csv.reader(io.StringIO(sample.text)))

    from openpyxl import Workbook, load_workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Team Name", "Leader Name", "Leader Email", "Leader Password"])
    sheet.append(["Excel Team", "Excel Leader", "excel.leader@example.com", "Excel@123"])
    source = io.BytesIO()
    workbook.save(source)
    workbook.close()

    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={
            "file": (
                "participants.xlsx",
                source.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200, imported.text
    immediate = client.get(
        f"/admin/registration/import/download/{imported.json()['download_token']}",
        headers=admin_headers,
    )
    output = load_workbook(io.BytesIO(immediate.content), data_only=True)
    assert output.active["D2"].value == "Excel@123"
    output.close()
    assert _login(client, "excel.leader@example.com", "Excel@123").status_code == 200
