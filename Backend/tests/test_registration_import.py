"""Registration import: teams/members/leaders, idempotency, credentials."""
import csv
import io
import logging
import re
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook, load_workbook
import pandas as pd
import pytest
from sqlalchemy import event

from app.core.config import settings
from app.core.security import get_password_hash, verify_password
from app.models.models import Bid, GameConfig, Member, ProblemStatement, RegistrationImport, RegistrationImportRow, RoundControl, Submission, Team, User, WalletTransaction, Wildcard, WildcardBid
from app.services.demo_seed import provision_demo_accounts


@pytest.fixture
def registration_log_capture(caplog):
    uvicorn_logger = logging.getLogger("uvicorn.error")
    previous_propagate = uvicorn_logger.propagate
    previous_disabled = uvicorn_logger.disabled
    uvicorn_logger.disabled = False
    uvicorn_logger.propagate = True
    caplog.set_level(logging.INFO, logger="uvicorn.error")
    try:
        yield caplog
    finally:
        uvicorn_logger.propagate = previous_propagate
        uvicorn_logger.disabled = previous_disabled


def test_import_preview_detects_teams_and_leaders(client, admin_headers, csv_bytes):
    response = client.post(
        "/admin/registration/import/preview",
        headers=admin_headers,
        files={"file": ("registrations.csv", csv_bytes, "text/csv")},
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["teams_detected"] == 3
    assert data["leaders_detected"] == 3
    assert data["members_detected"] == 4  # Aarav+Diya (alpha), Charlie+Rohan (beta)



def test_import_preview_does_not_persist_data(client, admin_headers, csv_bytes, db):
    response = client.post(
        "/admin/registration/import/preview",
        headers=admin_headers,
        files={"file": ("registrations.csv", csv_bytes, "text/csv")},
    )
    assert response.status_code == 200
    assert db.query(Team).count() == 0
    assert db.query(User).count() == 1  # only admin


def test_confirm_import_creates_teams_members_leader_and_wallets(client, admin_headers, csv_bytes, db):
    preview = client.post(
        "/admin/registration/import/preview",
        headers=admin_headers,
        files={"file": ("registrations.csv", csv_bytes, "text/csv")},
    ).json()
    import_id = preview["import_id"]

    response = client.post("/admin/registration/import/confirm", headers=admin_headers, json={"import_id": import_id})
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["teams_created"] == 3
    assert data["accounts_created"] == 7  # 3 leaders + all 4 teammates

    team_alpha = db.query(Team).filter(Team.team_name == "Team Alpha").first()
    assert team_alpha is not None
    assert team_alpha.is_approved is True
    assert team_alpha.coins == 5000  # EventConfig.starting_coins

    leader_email = db.query(User).filter(User.email == "alice@test.com").first()
    assert leader_email is not None
    assert leader_email.role == "leader"
    assert team_alpha.leader_id == leader_email.id

    member = db.query(Member).filter(Member.team_id == team_alpha.id, Member.member_name == "Aarav").first()
    assert member is not None
    assert member.email == "aarav@test.com"

    # member account created for Aarav
    aarav_user = db.query(User).filter(User.email == "aarav@test.com").first()
    assert aarav_user is not None
    assert aarav_user.role == "member"
    assert aarav_user.team_id == team_alpha.id

    # teammates without email receive a stable readable participant ID
    diya_user = db.query(User).filter(User.email == f"BTB-T{team_alpha.id:03d}-M02").first()
    assert diya_user is not None
    assert diya_user.name == "Diya"
    assert diya_user.team_id == team_alpha.id

    # wallet ledger entry
    tx = db.query(WalletTransaction).filter(WalletTransaction.team_id == team_alpha.id).first()
    assert tx is not None
    assert tx.transaction_type == "INITIAL_ALLOCATION"
    assert tx.amount == 5000


def test_duplicate_import_does_not_create_duplicates(client, admin_headers, csv_bytes, db):
    preview = client.post(
        "/admin/registration/import/preview",
        headers=admin_headers,
        files={"file": ("registrations.csv", csv_bytes, "text/csv")},
    ).json()
    first = client.post("/admin/registration/import/confirm", headers=admin_headers, json={"import_id": preview["import_id"]})
    assert first.status_code == 200

    # second import: same file uploaded again
    preview2 = client.post(
        "/admin/registration/import/preview",
        headers=admin_headers,
        files={"file": ("registrations.csv", csv_bytes, "text/csv")},
    ).json()
    assert all(row["status"] in ("update", "duplicate") for row in preview2["rows"])

    second = client.post("/admin/registration/import/confirm", headers=admin_headers, json={"import_id": preview2["import_id"]})
    assert second.status_code == 200
    assert second.json()["teams_created"] == 0
    assert second.json()["accounts_created"] == 0

    assert db.query(Team).count() == 3
    assert db.query(User).filter(User.role == "leader").count() == 3


def test_import_rejects_unsupported_file(client, admin_headers):
    response = client.post(
        "/admin/registration/import/preview",
        headers=admin_headers,
        files={"file": ("data.txt", b"not a spreadsheet", "text/plain")},
    )
    assert response.status_code == 400


def test_import_credential_export(client, admin_headers, csv_bytes):
    credentialed_csv = (
        "Team Name,Leader Name,Leader Email,Leader Password\n"
        "Team Alpha,Alice,alice@test.com,Alice@123\n"
        "Team Beta,Bob,bob@test.com,Bob@1234\n"
        "Team Gamma,Carol,carol@test.com,Carol@123\n"
    ).encode()
    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.csv", credentialed_csv, "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    exported = client.get(
        f"/admin/registration/import/download/{imported.json()['download_token']}",
        headers=admin_headers,
    )
    assert exported.status_code == 200
    rows = list(csv.DictReader(io.StringIO(exported.content.decode("utf-8-sig"))))
    assert len(rows) == 3
    assert {row["Leader Login Email"] for row in rows} == {
        "alice@test.com", "bob@test.com", "carol@test.com",
    }
    assert all(row["Leader Password"] for row in rows)
    assert all(row["Credential Status"] == "PASSWORD SET" for row in rows)


def test_import_requires_admin(client, csv_bytes):
    response = client.post(
        "/admin/registration/import/preview",
        files={"file": ("registrations.csv", csv_bytes, "text/csv")},
    )
    assert response.status_code in (401, 403)


def _registration_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Registrations"
    sheet.append([
        "Team Name", "Leader Name", "Leader Email", "Leader Password",
        "Member 1 Name", "Member 1 Email", "Member 1 Password",
        "Member 2 Name", "Member 2 Email", "Member 2 Password", "Organizer Notes",
    ])
    sheet.append(["Team Alpha", "Leader Alpha", "alpha@example.com", "AlphaLeader@123", "A One", "a.one@example.com", "AlphaMember1@123", "A Two", "a.two@example.com", "AlphaMember2@123", "Keep alpha note"])
    sheet.append(["Team Beta", "Leader Beta", "beta@example.com", "BetaLeader@123", "B One", "b.one@example.com", "BetaMember1@123", "B Two", "b.two@example.com", "BetaMember2@123", "Keep beta note"])
    sheet["A1"].font = sheet["A1"].font.copy(bold=True)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_xlsx_import_preserves_supplied_credentials(client, admin_headers, db):
    source = _registration_xlsx()
    response = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.xlsx", source, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200, response.text
    summary = response.json()
    assert summary == {
        **summary,
        "teams_processed": 2,
        "teams_created": 2,
        "teams_updated": 0,
        "leaders_created": 2,
        "existing_leaders": 0,
        "members_imported": 4,
        "rows_failed": 0,
        "errors": [],
    }

    download = client.get(
        f"/admin/registration/import/download/{summary['download_token']}",
        headers=admin_headers,
    )
    assert download.status_code == 200
    assert download.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    workbook = load_workbook(BytesIO(download.content), data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    assert headers == [
        "Team Name", "Leader Name", "Leader Email", "Leader Password",
        "Member 1 Name", "Member 1 Email", "Member 1 Password",
        "Member 2 Name", "Member 2 Email", "Member 2 Password", "Organizer Notes",
        "Leader Login Email", "Credential Status",
    ]
    assert sheet["K2"].value == "Keep alpha note"
    assert sheet["K3"].value == "Keep beta note"
    assert sheet["L2"].value == "alpha@example.com"
    assert sheet["L3"].value == "beta@example.com"
    assert sheet["M2"].value == "PASSWORD SET"
    assert sheet["M3"].value == "PASSWORD SET"
    alpha_password = sheet["D2"].value
    workbook.close()

    alpha = db.query(User).filter(User.email == "alpha@example.com").one()
    assert alpha.role == "leader"
    assert verify_password(alpha_password, alpha.password_hash)
    assert alpha.password_hash != alpha_password
    assert client.post("/login", data={"username": alpha.email, "password": alpha_password}).status_code == 200
    assert db.query(Member).count() == 4
    assert db.query(User).filter(User.role == "member").count() == 4
    assert client.get(
        f"/admin/registration/import/download/{summary['download_token']}",
        headers=admin_headers,
    ).status_code == 404

    second = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.xlsx", source, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    assert second["teams_created"] == 0
    assert second["teams_updated"] == 2
    assert second["leaders_created"] == 0
    assert second["existing_leaders"] == 2
    assert db.query(Team).count() == 2
    assert db.query(User).filter(User.role == "leader").count() == 2
    assert verify_password(
        alpha_password,
        db.query(User).filter(User.email == "alpha@example.com").one().password_hash,
    )

    second_download = client.get(
        f"/admin/registration/import/download/{second['download_token']}",
        headers=admin_headers,
    )
    workbook = load_workbook(BytesIO(second_download.content), data_only=True)
    assert workbook.active["M2"].value == "PASSWORD SET"
    assert workbook.active["M3"].value == "PASSWORD SET"
    workbook.close()


def test_registration_import_reports_duplicate_rows_and_requires_admin(client, admin_headers):
    csv_content = (
        "Team Name,Leader Name,Leader Email,Leader Password,Member 1 Name,Member 1 Email,Member 1 Password\n"
        "Team Alpha,Alpha,alpha@example.com,Alpha@123,Member A,member@example.com,MemberA@123\n"
        "Team Alpha,Beta,beta@example.com,Beta@123,Member B,other@example.com,MemberB@123\n"
        "Team Gamma,Gamma,gamma@example.com,Gamma@123,Member C,member@example.com,MemberC@123\n"
    ).encode()
    response = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.csv", csv_content, "text/csv")},
    )
    assert response.status_code == 200
    summary = response.json()
    assert summary["teams_processed"] == 1
    assert summary["rows_failed"] == 2
    assert any("already appears" in error["message"] for error in summary["errors"])
    assert any("already used" in error["message"] for error in summary["errors"])

    unauthorized = client.post(
        "/admin/registration/import",
        files={"file": ("registrations.csv", csv_content, "text/csv")},
    )
    assert unauthorized.status_code == 401


def test_direct_registration_import_reuses_preloaded_records_and_commits_once(
    client, admin_headers, db, engine, registration_log_capture
):
    caplog = registration_log_capture
    initial_source = (
        "Team Name,Leader Name,Leader Email,Leader Password,Member 1 Name,Member 1 Email,Member 1 Password,Member 2 Name,Member 2 Email,Member 2 Password\n"
        "Team Alpha,Leader Alpha,alpha@example.com,Alpha@123,Member Alpha,member.alpha@example.com,MemberAlpha@123,Member Without Email,,\n"
        "Team Beta,Leader Beta,beta@example.com,Beta@123,Member Beta,member.beta@example.com,MemberBeta@123,,,\n"
    ).encode()
    initial = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("initial.csv", initial_source, "text/csv")},
    )
    assert initial.status_code == 200, initial.text
    assert initial.json()["teams_created"] == 2, initial.json()
    alpha_hash = db.query(User).filter(User.email == "alpha@example.com").one().password_hash

    repeated_source = (
        "Team Name,Leader Name,Leader Email,Leader Password,Member 1 Name,Member 1 Email,Member 1 Password,Member 2 Name,Member 2 Email,Member 2 Password\n"
        "Team Alpha,Leader Alpha Updated,alpha@example.com,,Member Alpha Updated,member.alpha@example.com,,Member Without Email Updated,,,\n"
        "Team Beta,Leader Beta,beta@example.com,,Member Beta,member.beta@example.com,,,,\n"
        "Team Gamma,Leader Gamma,gamma@example.com,Gamma@123,Member Gamma,member.gamma@example.com,MemberGamma@123,,,\n"
    ).encode()

    selected_tables = {"teams": 0, "users": 0, "members": 0}
    commit_count = 0

    def count_selects(_conn, _cursor, statement, _parameters, _context, _executemany):
        if not statement.lstrip().upper().startswith("SELECT"):
            return
        for table_name in selected_tables:
            if re.search(rf"\bFROM\s+{table_name}\b", statement, re.IGNORECASE):
                selected_tables[table_name] += 1

    def count_commits(_conn):
        nonlocal commit_count
        commit_count += 1

    event.listen(engine, "before_cursor_execute", count_selects)
    event.listen(engine, "commit", count_commits)
    try:
        repeated = client.post(
            "/admin/registration/import",
            headers=admin_headers,
            files={"file": ("repeated.csv", repeated_source, "text/csv")},
        )
    finally:
        event.remove(engine, "before_cursor_execute", count_selects)
        event.remove(engine, "commit", count_commits)

    assert repeated.status_code == 200, repeated.text
    summary = repeated.json()
    assert summary["teams_created"] == 1
    assert summary["teams_updated"] == 2
    assert summary["leaders_created"] == 1
    assert summary["existing_leaders"] == 2
    assert summary["members_imported"] == 4
    assert summary["rows_failed"] == 0
    assert selected_tables == {"teams": 1, "users": 2, "members": 1}
    assert commit_count == 1

    assert db.query(Team).count() == 3
    assert db.query(User).filter(User.role == "leader").count() == 3
    assert db.query(User).filter(User.role == "member").count() == 4
    assert db.query(Member).count() == 4
    alpha = db.query(User).filter(User.email == "alpha@example.com").one()
    assert alpha.name == "Leader Alpha Updated"
    assert alpha.password_hash == alpha_hash
    assert db.query(User).filter(User.email == "member.alpha@example.com").count() == 1
    assert db.query(User).filter(User.email.like("BTB-T%-M02%")).count() == 1
    assert db.query(RegistrationImport).count() == 2
    assert db.query(RegistrationImportRow).count() == 5

    credentials = client.get(
        f"/admin/registration/import/download/{summary['download_token']}",
        headers=admin_headers,
    )
    assert credentials.status_code == 200
    assert credentials.headers["content-type"].startswith("text/csv")

    duplicate_source = (
        "Team Name,Leader Name,Leader Email,Leader Password,Member 1 Name,Member 1 Email,Member 1 Password\n"
        "Team Delta,Leader Delta,delta@example.com,Delta@123,Wrong Team Member,member.alpha@example.com,\n"
    ).encode()
    duplicate = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("duplicate.csv", duplicate_source, "text/csv")},
    )
    assert duplicate.status_code == 200, duplicate.text
    assert duplicate.json()["teams_processed"] == 0
    assert any(
        "belongs to another account or team" in error["message"]
        for error in duplicate.json()["errors"]
    )
    assert db.query(User).filter(User.email == "delta@example.com").count() == 0

    password_update_source = (
        "Team Name,Leader Name,Leader Email,Leader Password,Member 1 Name,Member 1 Email,Member 1 Password\n"
        "Team Alpha,Leader Alpha Updated,alpha@example.com,AlphaNew@123,Member Alpha Updated,member.alpha@example.com,\n"
    ).encode()
    password_update = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("password-update.csv", password_update_source, "text/csv")},
    )
    assert password_update.status_code == 200, password_update.text
    db.expire_all()
    updated_alpha = db.query(User).filter(User.email == "alpha@example.com").one()
    assert updated_alpha.password_hash != alpha_hash
    assert verify_password("AlphaNew@123", updated_alpha.password_hash)

    timing_messages = [record.getMessage() for record in caplog.records if "Registration import" in record.getMessage()]
    for stage in (
        "file_read", "parse_total", "pandas_parse", "column_normalization",
        "validation", "db_preload", "server_password_hashing", "db_processing",
        "credential_export", "commit", "total",
    ):
        assert any(f"stage={stage}" in message for message in timing_messages)
    assert any("rows=3 valid=3" in message for message in timing_messages)


def _registration_frame_bytes(frame: pd.DataFrame, suffix: str) -> bytes:
    if suffix == ".csv":
        return frame.to_csv(index=False).encode("utf-8")
    output = BytesIO()
    frame.to_excel(output, index=False, engine="openpyxl")
    return output.getvalue()


@pytest.mark.parametrize("suffix", [".csv", ".xlsx"])
def test_pandas_registration_imports_plaintext_csv_and_xlsx(
    suffix, client, admin_headers, db
):
    frame = pd.DataFrame([{
        "Team Name": f"Pandas Team {suffix}",
        "Leader Name": "Pandas Leader",
        "Leader Email": f"leader-{suffix[1:]}@pandas.test",
        "Leader Password": "LeaderPlain@123",
        "Member 1 Name": "Pandas Member",
        "Member 1 Email": f"member-{suffix[1:]}@pandas.test",
        "Member 1 Password": "MemberPlain@123",
    }])
    response = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": (f"registration{suffix}", _registration_frame_bytes(frame, suffix))},
    )
    assert response.status_code == 200, response.text
    assert response.json()["teams_created"] == 1
    assert response.json()["leaders_created"] == 1
    assert response.json()["members_imported"] == 1
    leader = db.query(User).filter(User.email == f"leader-{suffix[1:]}@pandas.test").one()
    member = db.query(User).filter(User.email == f"member-{suffix[1:]}@pandas.test").one()
    assert verify_password("LeaderPlain@123", leader.password_hash)
    assert verify_password("MemberPlain@123", member.password_hash)
    assert leader.team_id == member.team_id


def test_hash_columns_take_priority_and_never_export_hashes(client, admin_headers, db, registration_log_capture):
    caplog = registration_log_capture
    passwords = {
        "hash_only_leader": "HashOnlyLeader@123",
        "hash_only_member": "HashOnlyMember@123",
        "priority_leader": "PriorityLeader@123",
        "priority_member": "PriorityMember@123",
        "mixed_member": "MixedMember@123",
    }
    hashes = {name: get_password_hash(password) for name, password in passwords.items()}
    frame = pd.DataFrame([
        {
            "Team Name": "Hash Only Team",
            "Leader Name": "Hash Only Leader",
            "Leader Email": "hash-only-leader@test.com",
            "Leader Password Hash": hashes["hash_only_leader"],
            "Member 1 Name": "Hash Only Member",
            "Member 1 Email": "hash-only-member@test.com",
            "Member 1 Password Hash": hashes["hash_only_member"],
        },
        {
            "Team Name": "Hash Priority Team",
            "Leader Name": "Hash Priority Leader",
            "Leader Email": "hash-priority-leader@test.com",
            "Leader Password": "WrongLeaderPlaintext@123",
            "Leader Password Hash": hashes["priority_leader"],
            "Member 1 Name": "Hash Priority Member",
            "Member 1 Email": "hash-priority-member@test.com",
            "Member 1 Password": "WrongMemberPlaintext@123",
            "Member 1 Password Hash": hashes["priority_member"],
        },
        {
            "Team Name": "Mixed Credential Team",
            "Leader Name": "Mixed Leader",
            "Leader Email": "mixed-leader@test.com",
            "Leader Password": "MixedLeaderPlain@123",
            "Leader Password Hash": "",
            "Member 1 Name": "Mixed Member",
            "Member 1 Email": "mixed-member@test.com",
            "Member 1 Password Hash": hashes["mixed_member"],
        },
    ])
    source = _registration_frame_bytes(frame, ".xlsx")
    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("hashes.xlsx", source, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["teams_created"] == 3
    assert imported.json()["participant_accounts_created"] == 3

    expected_passwords = {
        "hash-only-leader@test.com": passwords["hash_only_leader"],
        "hash-only-member@test.com": passwords["hash_only_member"],
        "hash-priority-leader@test.com": passwords["priority_leader"],
        "hash-priority-member@test.com": passwords["priority_member"],
        "mixed-leader@test.com": "MixedLeaderPlain@123",
        "mixed-member@test.com": passwords["mixed_member"],
    }
    for email, password in expected_passwords.items():
        account = db.query(User).filter(User.email == email).one()
        assert account.credentials_active is True
        assert verify_password(password, account.password_hash)
    assert not verify_password(
        "WrongLeaderPlaintext@123",
        db.query(User).filter(User.email == "hash-priority-leader@test.com").one().password_hash,
    )
    assert not verify_password(
        "WrongMemberPlaintext@123",
        db.query(User).filter(User.email == "hash-priority-member@test.com").one().password_hash,
    )
    assert client.post(
        "/login",
        data={"username": "hash-only-leader@test.com", "password": passwords["hash_only_leader"]},
    ).status_code == 200
    assert client.post(
        "/login",
        data={"username": "hash-only-member@test.com", "password": passwords["hash_only_member"]},
    ).status_code == 200

    completion_messages = [
        record.getMessage() for record in caplog.records
        if record.getMessage().startswith("Registration import completed")
    ]
    assert any("supplied_hashes=5 server_hashes=1" in message for message in completion_messages)
    stored_rows = db.query(RegistrationImportRow).all()
    assert all("password_hash" not in row.members_json for row in stored_rows)
    assert all("$2" not in row.source_values_json for row in stored_rows)

    credentials = client.get(
        f"/admin/registration/import/download/{imported.json()['download_token']}",
        headers=admin_headers,
    )
    workbook = load_workbook(BytesIO(credentials.content), data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    hash_columns = [index + 1 for index, header in enumerate(headers) if "Password Hash" in str(header)]
    assert hash_columns
    assert all(
        sheet.cell(row=row, column=column).value in (None, "NOT EXPORTED")
        for row in range(2, sheet.max_row + 1)
        for column in hash_columns
    )
    workbook.close()

    assignment = client.get("/admin/registration/assignments", headers=admin_headers)
    assert assignment.status_code == 200, assignment.text
    workbook = load_workbook(BytesIO(assignment.content), data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    hash_columns = [index + 1 for index, header in enumerate(headers) if "Password Hash" in str(header)]
    assert all(
        sheet.cell(row=row, column=column).value in (None, "NOT EXPORTED")
        for row in range(2, sheet.max_row + 1)
        for column in hash_columns
    )
    workbook.close()


def test_non_sha256_hash_is_rejected_without_fallback(client, admin_headers):
    frame = pd.DataFrame([{
        "Team Name": "Invalid Hash Team",
        "Leader Name": "Invalid Hash Leader",
        "Leader Email": "invalid-hash@test.com",
        "Leader Password": "ValidPlaintext@123",
        "Leader Password Hash": "$2b$12$not-a-valid-bcrypt-hash",
    }])
    response = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("invalid-hash.csv", _registration_frame_bytes(frame, ".csv"), "text/csv")},
    )
    assert response.status_code == 200, response.text
    assert response.json()["teams_processed"] == 0
    assert any("structurally valid sha256$salt$digest hash" in error["message"] for error in response.json()["errors"])


def test_eighty_team_hash_import_avoids_server_hashing_and_repeated_selects(
    client, admin_headers, engine, registration_log_capture
):
    caplog = registration_log_capture
    shared_hash = get_password_hash("BenchmarkHash@123")
    rows = []
    for number in range(1, 81):
        rows.append({
            "Team Name": f"Benchmark Team {number:02d}",
            "Leader Name": f"Benchmark Leader {number:02d}",
            "Leader Email": f"leader{number:02d}@benchmark.test",
            "Leader Password Hash": shared_hash,
            "Member 1 Name": f"Benchmark Member A {number:02d}",
            "Member 1 Email": f"member-a-{number:02d}@benchmark.test",
            "Member 1 Password Hash": shared_hash,
            "Member 2 Name": f"Benchmark Member B {number:02d}",
            "Member 2 Email": f"member-b-{number:02d}@benchmark.test",
            "Member 2 Password Hash": shared_hash,
        })

    selected_tables = {"teams": 0, "users": 0, "members": 0}

    def count_selects(_conn, _cursor, statement, _parameters, _context, _executemany):
        if not statement.lstrip().upper().startswith("SELECT"):
            return
        for table_name in selected_tables:
            if re.search(rf"\bFROM\s+{table_name}\b", statement, re.IGNORECASE):
                selected_tables[table_name] += 1

    event.listen(engine, "before_cursor_execute", count_selects)
    try:
        imported = client.post(
            "/admin/registration/import",
            headers=admin_headers,
            files={
                "file": (
                    "benchmark.csv",
                    _registration_frame_bytes(pd.DataFrame(rows), ".csv"),
                    "text/csv",
                )
            },
        )
    finally:
        event.remove(engine, "before_cursor_execute", count_selects)

    assert imported.status_code == 200, imported.text
    assert imported.json()["teams_created"] == 80
    assert imported.json()["leaders_created"] == 80
    assert imported.json()["participant_accounts_created"] == 160
    assert selected_tables == {"teams": 1, "users": 2, "members": 1}
    completion_messages = [
        record.getMessage() for record in caplog.records
        if record.getMessage().startswith("Registration import completed")
    ]
    assert any("rows=80 valid=80" in message for message in completion_messages)
    assert any("supplied_hashes=240 server_hashes=0" in message for message in completion_messages)


def test_demo_csv_is_importer_compatible_and_returns_csv_credentials(client, admin_headers, db):
    demo = client.get("/admin/registration/demo.csv", headers=admin_headers)
    assert demo.status_code == 200
    source_rows = list(csv.reader(io.StringIO(demo.content.decode("utf-8-sig"))))
    assert source_rows[0] == [
        "Team Name", "Leader Name", "Leader Email", "Leader Password",
        "Member 1 Name", "Member 1 Email", "Member 1 Password",
        "Member 2 Name", "Member 2 Email", "Member 2 Password",
    ]

    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("demo-registration.csv", demo.content, "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    summary = imported.json()
    assert summary["teams_created"] == 2
    assert summary["leaders_created"] == 2
    assert summary["download_filename"].endswith(".csv")

    downloaded = client.get(
        f"/admin/registration/import/download/{summary['download_token']}",
        headers=admin_headers,
    )
    assert downloaded.status_code == 200
    assert downloaded.headers["content-type"].startswith("text/csv")
    output_rows = list(csv.reader(io.StringIO(downloaded.content.decode("utf-8-sig"))))
    assert output_rows[0] == [*source_rows[0], "Leader Login Email", "Credential Status"]
    assert [row[:10] for row in output_rows[1:]] == source_rows[1:]
    assert output_rows[1][-2] == "alice@example.com"
    assert output_rows[1][-1] == "PASSWORD SET"
    alice_password = source_rows[1][3]
    alice = db.query(User).filter(User.email == "alice@example.com").one()
    assert verify_password(alice_password, alice.password_hash)
    repeated = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("demo-registration.csv", demo.content, "text/csv")},
    ).json()
    assert repeated["teams_created"] == 0
    assert repeated["leaders_created"] == 0
    assert db.query(User).filter(User.email == "alice@example.com").count() == 1
    assert verify_password(
        alice_password,
        db.query(User).filter(User.email == "alice@example.com").one().password_hash,
    )
    repeated_download = client.get(
        f"/admin/registration/import/download/{repeated['download_token']}",
        headers=admin_headers,
    )
    repeated_rows = list(csv.reader(io.StringIO(repeated_download.content.decode("utf-8-sig"))))
    assert repeated_rows[1][-1] == "PASSWORD SET"


def test_registration_credential_reset_allows_fresh_passwords_and_preserves_system_accounts(client, admin_headers, db):
    provision_demo_accounts(db)
    manual_leader = User(
        name="Manual Leader",
        email="manual.leader@example.com",
        password_hash=get_password_hash("ManualLeader@123"),
        role="leader",
    )
    db.add(manual_leader)
    db.flush()
    manual_team = Team(team_name="Manual Team", leader_id=manual_leader.id, is_approved=True)
    db.add(manual_team)
    db.flush()
    manual_leader.team_id = manual_team.id
    db.add_all([
        User(name="Manual Admin", email="manual.admin@example.com", password_hash=get_password_hash("ManualAdmin@123"), role="admin"),
        User(name="Manual Display", email="manual.display@example.com", password_hash=get_password_hash("ManualDisplay@123"), role="display"),
    ])
    db.commit()
    source = (
        "Team Name,Leader Name,Leader Email,Leader Password,Member 1 Name,Member 1 Email,Member 1 Password\n"
        "Team Alpha,Alpha Leader,alpha@example.com,Alpha@123,Alpha Member,member.alpha@example.com,AlphaMember@123\n"
        "Team Beta,Beta Leader,beta@example.com,Beta@123,Beta Member,member.beta@example.com,BetaMember@123\n"
    ).encode()

    def import_and_download():
        imported = client.post(
            "/admin/registration/import",
            headers=admin_headers,
            files={"file": ("registrations.csv", source, "text/csv")},
        )
        assert imported.status_code == 200, imported.text
        downloaded = client.get(
            f"/admin/registration/import/download/{imported.json()['download_token']}",
            headers=admin_headers,
        )
        assert downloaded.status_code == 200, downloaded.text
        rows = list(csv.DictReader(io.StringIO(downloaded.content.decode("utf-8-sig"))))
        return imported.json(), rows

    first, first_rows = import_and_download()
    assert first["leaders_created"] == 2
    first_passwords = {row["Leader Login Email"]: row["Leader Password"] for row in first_rows}
    assert all(password and password != "EXISTING ACCOUNT" for password in first_passwords.values())
    for email, password in first_passwords.items():
        account = db.query(User).filter(User.email == email).one()
        assert account.password_hash != password
        assert verify_password(password, account.password_hash)
        assert client.post("/login", data={"username": email, "password": password}).status_code == 200

    repeated, repeated_rows = import_and_download()
    assert repeated["leaders_created"] == 0
    assert repeated["existing_leaders"] == 2
    assert {row["Leader Login Email"] for row in repeated_rows} == {"alpha@example.com", "beta@example.com"}
    assert db.query(User).filter(User.email.in_(("alpha@example.com", "beta@example.com"))).count() == 2

    alpha_headers = {
        "Authorization": "Bearer " + client.post(
            "/login",
            data={"username": "alpha@example.com", "password": first_passwords["alpha@example.com"]},
        ).json()["access_token"]
    }
    assert client.post(
        "/admin/registration/credentials/reset",
        json={"confirmation": "RESET CREDENTIALS"},
    ).status_code == 401
    assert client.post(
        "/admin/registration/credentials/reset",
        headers=admin_headers,
        json={"confirmation": "wrong"},
    ).status_code == 422
    assert client.post(
        "/admin/registration/credentials/reset",
        headers=alpha_headers,
        json={"confirmation": "RESET CREDENTIALS"},
    ).status_code == 403

    reset = client.post(
        "/admin/registration/credentials/reset",
        headers=admin_headers,
        json={"confirmation": "RESET CREDENTIALS"},
    )
    assert reset.status_code == 200, reset.text
    assert reset.json()["reset"]["participant_accounts"] == 4
    assert reset.json()["deleted"]["participant_accounts"] == 4
    assert reset.json()["deleted"]["teams"] == 2
    assert reset.json()["event_lifecycle_reset"] is False
    db.expire_all()
    assert db.query(User).filter(User.account_source == "IMPORTED").count() == 0
    assert db.query(Team).filter(Team.team_name.in_(("Team Alpha", "Team Beta"))).count() == 0
    for email, password in first_passwords.items():
        assert client.post("/login", data={"username": email, "password": password}).status_code == 401
    assert client.get("/participant/dashboard", headers=alpha_headers).status_code == 401
    assert client.post("/login", data={"username": settings.DEMO_LEADER_EMAIL, "password": settings.DEMO_LEADER_PASSWORD}).status_code == 200
    assert client.post("/login", data={"username": settings.DEMO_ADMIN_EMAIL, "password": settings.DEMO_ADMIN_PASSWORD}).status_code == 200
    assert client.post(
        "/leaderboard/login",
        data={"username": settings.LEADERBOARD_DISPLAY_EMAIL, "password": settings.LEADERBOARD_DISPLAY_PASSWORD},
    ).status_code == 200
    assert client.post("/login", data={"username": "manual.admin@example.com", "password": "ManualAdmin@123"}).status_code == 200
    assert client.post("/login", data={"username": "manual.leader@example.com", "password": "ManualLeader@123"}).status_code == 200
    assert client.post(
        "/leaderboard/login",
        data={"username": "manual.display@example.com", "password": "ManualDisplay@123"},
    ).status_code == 200

    after_reset, after_reset_rows = import_and_download()
    assert after_reset["teams_created"] == 2
    assert after_reset["leaders_created"] == 2
    assert after_reset["existing_leaders"] == 0
    new_passwords = {row["Leader Login Email"]: row["Leader Password"] for row in after_reset_rows}
    assert all(password and password != "EXISTING ACCOUNT" for password in new_passwords.values())
    for email, password in new_passwords.items():
        assert client.post("/login", data={"username": email, "password": password}).status_code == 200


def test_registration_reset_preserves_manual_team_updated_by_import(client, admin_headers, db):
    manual_leader = User(
        name="Manual Leader",
        email="manual.team@example.com",
        password_hash=get_password_hash("ManualTeam@123"),
        role="leader",
    )
    db.add(manual_leader)
    db.flush()
    manual_team = Team(team_name="Manual Existing Team", leader_id=manual_leader.id, is_approved=True)
    db.add(manual_team)
    db.flush()
    manual_leader.team_id = manual_team.id
    db.commit()

    source = (
        "Team Name,Leader Name,Leader Email,Leader Password,Member 1 Name,Member 1 Email,Member 1 Password\n"
        "Manual Existing Team,Manual Leader,manual.team@example.com,,Imported Member,imported.member@example.com,Member@123\n"
    ).encode()
    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.csv", source, "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["teams_updated"] == 1

    reset = client.post(
        "/admin/registration/credentials/reset",
        headers=admin_headers,
        json={"confirmation": "RESET CREDENTIALS"},
    )
    assert reset.status_code == 200, reset.text
    assert reset.json()["deleted"]["teams"] == 0
    db.expire_all()
    assert db.query(Team).filter(Team.id == manual_team.id).count() == 1
    assert db.query(User).filter(User.id == manual_leader.id, User.account_source == "MANUAL").count() == 1
    assert db.query(User).filter(User.email == "imported.member@example.com").count() == 0
    assert db.query(Member).filter(Member.team_id == manual_team.id).count() == 0
    assert db.query(RegistrationImportRow).count() == 0
    assert client.post(
        "/login",
        data={"username": "manual.team@example.com", "password": "ManualTeam@123"},
    ).status_code == 200


def test_five_team_import_presence_reset_and_clean_reimport(client, admin_headers, db):
    rows = [
        f"Team {index},Leader {index},leader{index}@example.com,Leader{index}@123"
        for index in range(1, 6)
    ]
    source = ("Team Name,Leader Name,Leader Email,Leader Password\n" + "\n".join(rows) + "\n").encode()

    def import_sheet():
        response = client.post(
            "/admin/registration/import",
            headers=admin_headers,
            files={"file": ("five-teams.csv", source, "text/csv")},
        )
        assert response.status_code == 200, response.text
        return response.json()

    first_import = import_sheet()
    assert first_import["teams_created"] == 5
    initial_teams = client.get("/teams", headers=admin_headers).json()
    assert len(initial_teams) == 5
    assert all(team["logged_in"] is False for team in initial_teams)

    logins = [
        client.post(
            "/login",
            data={"username": f"leader{index}@example.com", "password": f"Leader{index}@123"},
        )
        for index in (1, 2)
    ]
    assert all(login.status_code == 200 for login in logins)
    first_token, second_token = (login.json()["access_token"] for login in logins)
    first_headers = {"Authorization": f"Bearer {first_token}"}

    with client.websocket_connect(f"/ws/auction?token={first_token}") as first_socket:
        assert first_socket.receive_json()["type"] == "event_snapshot"
        with client.websocket_connect(f"/ws/auction?token={second_token}") as second_socket:
            assert second_socket.receive_json()["type"] == "event_snapshot"
            active_teams = client.get("/teams", headers=admin_headers).json()
            assert sum(team["logged_in"] for team in active_teams) == 2

            reset = client.post(
                "/admin/registration/credentials/reset",
                headers=admin_headers,
                json={"confirmation": "RESET CREDENTIALS"},
            )
            assert reset.status_code == 200, reset.text
            assert reset.json()["deleted"]["teams"] == 5
            assert reset.json()["deleted"]["participant_accounts"] == 5
            assert reset.json()["event_lifecycle_reset"] is False

    assert client.get("/teams", headers=admin_headers).json() == []
    assert client.get("/participant/dashboard", headers=first_headers).status_code == 401
    assert db.query(RegistrationImportRow).count() == 0

    second_import = import_sheet()
    assert second_import["teams_created"] == 5
    assert second_import["teams_updated"] == 0
    reimported_teams = client.get("/teams", headers=admin_headers).json()
    assert len(reimported_teams) == 5
    assert all(team["logged_in"] is False for team in reimported_teams)

    post_reset_login = client.post(
        "/login",
        data={"username": "leader1@example.com", "password": "Leader1@123"},
    )
    assert post_reset_login.status_code == 200
    post_reset_token = post_reset_login.json()["access_token"]
    with client.websocket_connect(f"/ws/auction?token={post_reset_token}") as socket:
        assert socket.receive_json()["type"] == "event_snapshot"
        post_reset_teams = client.get("/teams", headers=admin_headers).json()
        assert sum(team["logged_in"] for team in post_reset_teams) == 1


def test_assignment_export_preserves_xlsx_format_and_original_columns(client, admin_headers):
    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.xlsx", _registration_xlsx(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200, imported.text
    exported = client.get("/admin/registration/assignments", headers=admin_headers)
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith("application/vnd.openxmlformats")
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    row = {header: sheet.cell(row=2, column=index + 1).value for index, header in enumerate(headers)}
    workbook.close()
    assert row["Organizer Notes"] == "Keep alpha note"
    assert row["Leader Login Email"] == "alpha@example.com"
    assert "Round 1 Problem Number" in headers
    assert "Wildcard Problem Description" in headers
    assert "Final Problem Title" in headers


def test_updated_registration_export_reflects_three_team_round_and_wildcard_assignments(client, admin_headers, db):
    source = (
        "Team Name,Leader Name,Leader Email,Leader Password,Organizer Notes\n"
        "Team A,Leader A,leader.a@example.com,LeaderA@123,Keep A\n"
        "Team B,Leader B,leader.b@example.com,LeaderB@123,Keep B\n"
        "Team C,Leader C,leader.c@example.com,LeaderC@123,Keep C\n"
    ).encode()
    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.csv", source, "text/csv")},
    )
    assert imported.status_code == 200, imported.text

    round_problems = [
        ProblemStatement(ps_number=f"R1-{index}", title=f"Round title {index}", description=f"Round description {index}", round=1, status="allocated")
        for index in range(1, 4)
    ]
    wildcard_problem = ProblemStatement(
        ps_number="WC-5", title="Wildcard title 5", description="Wildcard description 5", round=2, status="allocated",
    )
    db.add_all([*round_problems, wildcard_problem])
    db.flush()
    teams = {team.team_name: team for team in db.query(Team).filter(Team.team_name.in_(("Team A", "Team B", "Team C"))).all()}
    for index, team_name in enumerate(("Team A", "Team B", "Team C")):
        teams[team_name].round1_problem_id = round_problems[index].id
        teams[team_name].ps_id = round_problems[index].id
    teams["Team B"].wildcard_problem_id = wildcard_problem.id
    teams["Team B"].ps_id = wildcard_problem.id
    db.commit()

    exported = client.get("/admin/registration/assignments", headers=admin_headers)
    assert exported.status_code == 200, exported.text
    reader = csv.DictReader(io.StringIO(exported.content.decode("utf-8-sig")))
    rows = {row["Team Name"]: row for row in reader}
    assert reader.fieldnames[:5] == ["Team Name", "Leader Name", "Leader Email", "Leader Password", "Organizer Notes"]
    assert [rows[name]["Organizer Notes"] for name in ("Team A", "Team B", "Team C")] == ["Keep A", "Keep B", "Keep C"]

    assert rows["Team A"]["Round 1 Problem Number"] == "R1-1"
    assert rows["Team A"]["Round 1 Problem Title"] == "Round title 1"
    assert rows["Team A"]["Round 1 Problem Description"] == "Round description 1"
    assert rows["Team A"]["Wildcard Problem Number"] == ""
    assert rows["Team A"]["Final Problem Number"] == "R1-1"

    assert rows["Team B"]["Round 1 Problem Number"] == "R1-2"
    assert rows["Team B"]["Round 1 Problem Title"] == "Round title 2"
    assert rows["Team B"]["Wildcard Problem Number"] == "WC-5"
    assert rows["Team B"]["Wildcard Problem Title"] == "Wildcard title 5"
    assert rows["Team B"]["Wildcard Problem Description"] == "Wildcard description 5"
    assert rows["Team B"]["Final Problem Number"] == "WC-5"

    assert rows["Team C"]["Round 1 Problem Number"] == "R1-3"
    assert rows["Team C"]["Round 1 Problem Title"] == "Round title 3"
    assert rows["Team C"]["Round 1 Problem Description"] == "Round description 3"
    assert rows["Team C"]["Wildcard Problem Number"] == ""
    assert rows["Team C"]["Final Problem Number"] == "R1-3"


def test_assignment_export_never_replays_uploaded_plaintext_password(client, admin_headers):
    source = (
        "Team Name,Leader Name,Leader Email,Leader Password\n"
        "Secure Team,Secure Leader,secure@example.com,DoNotReplay123!\n"
    ).encode()
    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.csv", source, "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    exported = client.get("/admin/registration/assignments", headers=admin_headers)
    row = next(csv.DictReader(io.StringIO(exported.content.decode("utf-8-sig"))))
    assert row["Leader Password"] == "NOT EXPORTED"
    assert "DoNotReplay123!" not in exported.text


@pytest.mark.parametrize("active_state", ["ROUND1_BIDDING", "WILDCARD_BIDDING", "SUBMISSION"])
def test_registration_credential_reset_preserves_active_event_data(client, admin_headers, db, active_state):
    provision_demo_accounts(db)
    db.commit()
    source = (
        "Team Name,Leader Name,Leader Email,Leader Password,Member 1 Name,Member 1 Email,Member 1 Password\n"
        "Team Alpha,Alpha Leader,alpha@example.com,Alpha@123,Alpha Member,member.alpha@example.com,Member@123\n"
    ).encode()
    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.csv", source, "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    team = db.query(Team).filter(Team.team_name == "Team Alpha").one()
    imported_team_id = team.id
    leader = db.query(User).filter(User.email == "alpha@example.com").one()
    round_problem = ProblemStatement(ps_number="RESET-PS", title="Active", description="Active", round=1)
    wildcard_problem = ProblemStatement(ps_number="RESET-WC", title="Wildcard", description="Wildcard", round=2)
    db.add_all([round_problem, wildcard_problem])
    db.flush()
    team.ps_id = round_problem.id
    team.round1_problem_id = round_problem.id
    db.add_all([
        Bid(team_id=team.id, ps_id=round_problem.id, amount=100, round=1),
        Wildcard(team_id=team.id, coins_paid=0, status="applied"),
        WildcardBid(team_id=team.id, amount=175),
        Submission(
            team_id=team.id,
            problem_id=round_problem.id,
            submitted_by_user_id=leader.id,
            repository_url="https://github.com/example/preserved",
        ),
    ])
    game = db.query(GameConfig).first()
    game.state = active_state
    timer_end = datetime.now(timezone.utc) + timedelta(minutes=5)
    game.auction_timer_end = timer_end
    round1_control = db.query(RoundControl).filter(RoundControl.round_type == "ROUND1").first()
    if round1_control is None:
        round1_control = RoundControl(round_type="ROUND1")
        db.add(round1_control)
    round1_control.status = "BIDDING"
    round1_control.current_problem_id = round_problem.id
    wildcard_control = db.query(RoundControl).filter(RoundControl.round_type == "WILDCARD").first()
    if wildcard_control is None:
        wildcard_control = RoundControl(round_type="WILDCARD")
        db.add(wildcard_control)
    wildcard_control.status = "PROBLEM_SELECTION"
    wildcard_control.current_selection_rank = 1
    wildcard_control.selection_started_at = datetime.now(timezone.utc)
    wildcard_control.selection_ends_at = datetime.now(timezone.utc) + timedelta(seconds=30)
    wildcard_control.selection_duration_seconds = 30
    db.commit()

    reset = client.post(
        "/admin/registration/credentials/reset",
        headers=admin_headers,
        json={"confirmation": "RESET CREDENTIALS"},
    )
    assert reset.status_code == 200, reset.text
    assert reset.json()["event_state"] == active_state
    assert reset.json()["reset"]["participant_accounts"] == 2
    assert reset.json()["deleted"]["teams"] == 1
    db.expire_all()
    assert db.query(User).filter(User.email == "alpha@example.com").count() == 0
    assert db.query(Team).filter(Team.team_name == "Team Alpha").count() == 0
    assert db.query(Bid).filter(Bid.team_id == imported_team_id).count() == 0
    assert db.query(Wildcard).filter(Wildcard.team_id == imported_team_id).count() == 0
    assert db.query(WildcardBid).filter(WildcardBid.team_id == imported_team_id).count() == 0
    assert db.query(Submission).filter(Submission.team_id == imported_team_id).count() == 0
    assert db.query(ProblemStatement).count() == 2
    preserved_game = db.query(GameConfig).first()
    assert preserved_game.state == active_state
    assert preserved_game.auction_timer_end == timer_end
    controls = {control.round_type: control for control in db.query(RoundControl).all()}
    assert controls["ROUND1"].status == "BIDDING"
    assert controls["ROUND1"].current_problem_id == round_problem.id
    assert controls["WILDCARD"].status == "PROBLEM_SELECTION"
    assert controls["WILDCARD"].current_selection_rank == 1
    assert controls["WILDCARD"].selection_started_at is not None
    assert controls["WILDCARD"].selection_ends_at is not None
    assert client.post("/login", data={"username": settings.DEMO_LEADER_EMAIL, "password": settings.DEMO_LEADER_PASSWORD}).status_code == 200
    assert client.post("/login", data={"username": settings.DEMO_ADMIN_EMAIL, "password": settings.DEMO_ADMIN_PASSWORD}).status_code == 200
    assert client.post(
        "/leaderboard/login",
        data={"username": settings.LEADERBOARD_DISPLAY_EMAIL, "password": settings.LEADERBOARD_DISPLAY_PASSWORD},
    ).status_code == 200

    imported_again = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.csv", source, "text/csv")},
    )
    assert imported_again.status_code == 200, imported_again.text
    assert imported_again.json()["teams_created"] == 1
    assert imported_again.json()["leaders_created"] == 1
    assert imported_again.json()["existing_leaders"] == 0
    credentials = client.get(
        f"/admin/registration/import/download/{imported_again.json()['download_token']}",
        headers=admin_headers,
    )
    row = next(csv.DictReader(io.StringIO(credentials.content.decode("utf-8-sig"))))
    assert row["Leader Password"] not in {"", "EXISTING ACCOUNT"}
    assert client.post(
        "/login",
        data={"username": row["Leader Login Email"], "password": row["Leader Password"]},
    ).status_code == 200
