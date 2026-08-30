from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.database import Base
from app.core.security import get_password_hash
from app.models.models import Bid, ProblemStatement, RoundControl, Team, User, WalletTransaction
from app.services.round1_assignment import Round1AssignmentError, change_round1_problem_assignment


def _problem(db, number: int) -> ProblemStatement:
    problem = ProblemStatement(
        ps_number=f"R1-{number}",
        title=f"Problem {number}",
        description=f"Complete description for problem {number}",
        round=1,
        status="completed",
    )
    db.add(problem)
    db.flush()
    return problem


def _team(db, name: str, problem: ProblemStatement | None = None, *, coins: int = 1000) -> Team:
    slug = name.lower().replace(" ", "-")
    leader = User(
        name=f"{name} Leader",
        email=f"{slug}@change.test",
        password_hash=get_password_hash("temp-pass"),
        role="leader",
    )
    db.add(leader)
    db.flush()
    team = Team(
        team_name=name,
        coins=coins,
        leader_id=leader.id,
        ps_id=problem.id if problem else None,
        round1_problem_id=problem.id if problem else None,
        round1_assignment_type="BID_WINNER" if problem else None,
        round1_assignment_cost=300 if problem else None,
        is_approved=True,
        is_system_team=False,
    )
    db.add(team)
    db.flush()
    leader.team_id = team.id
    return team


def _round_control(db, *, ended: bool = False) -> RoundControl:
    control = RoundControl(
        round_type="ROUND1",
        status="CLOSED" if ended else "READY",
        ended=ended,
        round1_winning_bid_sum=900,
        round1_winning_bid_count=3,
    )
    db.add(control)
    db.flush()
    return control


def _change(client, admin_headers, team: Team, target: ProblemStatement):
    return client.put(
        f"/admin/rounds/round-1/assignments/{team.id}",
        headers=admin_headers,
        json={"target_problem_id": target.id},
    )


def _exported_team(client, admin_headers, team_name: str) -> dict[str, object]:
    response = client.get("/admin/rounds/round-1/assignments/export", headers=admin_headers)
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    try:
        values = list(workbook["Participant Assignments"].iter_rows(values_only=True))
        headers = [str(value) for value in values[0]]
        return next(dict(zip(headers, row)) for row in values[1:] if row[0] == team_name)
    finally:
        workbook.close()


def test_normal_change_moves_current_assignment_without_balance_or_aggregate_change(client, admin_headers, db):
    old, target = _problem(db, 1), _problem(db, 2)
    team = _team(db, "Team Alpha", old, coins=1700)
    control = _round_control(db)
    db.commit()

    response = _change(client, admin_headers, team, target)

    assert response.status_code == 200, response.text
    db.expire_all()
    team = db.get(Team, team.id)
    control = db.get(RoundControl, control.id)
    assert team.round1_problem_id == target.id and team.ps_id == target.id
    assert team.coins == 1700
    assert (control.round1_winning_bid_sum, control.round1_winning_bid_count) == (900, 3)
    snapshot = response.json()
    old_row = next(row for row in snapshot["problems"] if row["id"] == old.id)
    target_row = next(row for row in snapshot["problems"] if row["id"] == target.id)
    assert old_row["assigned_team_count"] == 0
    assert target_row["assigned_team_count"] == 1
    exported = _exported_team(client, admin_headers, "Team Alpha")
    assert exported["Round 1 Problem Number"] == "R1-2"
    assert exported["Round 1 Final Price / Winning Bid"] == 300


def test_full_target_rejects_change_and_preserves_original_assignment(client, admin_headers, db):
    old, target = _problem(db, 1), _problem(db, 3)
    moving = _team(db, "Moving Team", old)
    for index in range(5):
        _team(db, f"Target Team {index}", target)
    _round_control(db)
    db.commit()

    response = _change(client, admin_headers, moving, target)

    assert response.status_code == 409
    assert "full (5/5)" in response.json()["detail"]
    db.expire_all()
    moving = db.get(Team, moving.id)
    assert moving.round1_problem_id == old.id and moving.ps_id == old.id
    assert db.query(Team).filter(Team.round1_problem_id == target.id).count() == 5


def test_previously_unassigned_team_gets_problem_without_coin_deduction(client, admin_headers, db):
    target = _problem(db, 4)
    team = _team(db, "Team Beta", coins=1250)
    _round_control(db)
    db.commit()

    response = _change(client, admin_headers, team, target)

    assert response.status_code == 200, response.text
    db.expire_all()
    team = db.get(Team, team.id)
    assert team.round1_problem_id == target.id and team.ps_id == target.id
    assert team.round1_assignment_type == "MANUAL_ASSIGNMENT"
    assert team.round1_assignment_cost == 0 and team.coins == 1250
    assert db.query(WalletTransaction).filter(WalletTransaction.team_id == team.id).count() == 0
    exported = _exported_team(client, admin_headers, "Team Beta")
    assert exported["Round 1 Final Price / Winning Bid"] == 0


def test_participant_websocket_receives_authoritative_assignment_change(client, admin_headers, db):
    old, target = _problem(db, 1), _problem(db, 4)
    team = _team(db, "Realtime Team", old)
    _round_control(db)
    db.commit()
    login = client.post("/login", data={"username": "realtime-team@change.test", "password": "temp-pass"})
    assert login.status_code == 200, login.text

    with client.websocket_connect(f"/ws/auction?token={login.json()['access_token']}") as websocket:
        assert websocket.receive_json()["type"] == "event_snapshot"
        response = _change(client, admin_headers, team, target)
        assert response.status_code == 200, response.text
        event = websocket.receive_json()

    assert event["type"] == "round1_assignment_changed"
    assert event["payload"]["team_id"] == team.id
    assert event["payload"]["problem"]["id"] == target.id
    assert event["payload"]["problem"]["description"] == target.description


def test_participant_refresh_loads_changed_problem_from_database(client, admin_headers, db):
    old, target = _problem(db, 1), _problem(db, 4)
    team = _team(db, "Refresh Team", old)
    _round_control(db)
    db.commit()
    login = client.post("/login", data={"username": "refresh-team@change.test", "password": "temp-pass"})
    participant_headers = {"Authorization": f"Bearer {login.json()['access_token']}"}

    assert _change(client, admin_headers, team, target).status_code == 200
    dashboard = client.get("/participant/dashboard", headers=participant_headers)

    assert dashboard.status_code == 200, dashboard.text
    assert dashboard.json()["currentProblem"]["id"] == target.id
    assert dashboard.json()["round1Problem"]["id"] == target.id
    assert dashboard.json()["currentProblem"]["description"] == target.description


def test_unassigned_list_is_authoritative_and_updates_after_assignment(client, admin_headers, db):
    target = _problem(db, 3)
    assigned = _team(db, "Assigned Team", target)
    waiting = _team(db, "Waiting Team")
    _round_control(db)
    db.commit()

    before = client.get("/admin/rounds/round-1/assignments", headers=admin_headers)
    assert before.status_code == 200, before.text
    assert [row["team_id"] for row in before.json()["unassigned_teams"]] == [waiting.id]
    assert assigned.id not in [row["team_id"] for row in before.json()["unassigned_teams"]]

    changed = _change(client, admin_headers, waiting, target)
    assert changed.status_code == 200, changed.text
    assert changed.json()["unassigned_teams"] == []


def test_historical_bid_wallet_and_winning_average_survive_change(client, admin_headers, db):
    old, target = _problem(db, 1), _problem(db, 4)
    team = _team(db, "History Team", old, coins=700)
    control = _round_control(db)
    db.add(Bid(team_id=team.id, ps_id=old.id, amount=300, round=1))
    transaction = WalletTransaction(
        team_id=team.id,
        transaction_type="ROUND1_WIN",
        amount=-300,
        description="Round 1 win for problem 1",
    )
    db.add(transaction)
    db.commit()

    response = _change(client, admin_headers, team, target)

    assert response.status_code == 200, response.text
    db.expire_all()
    assert db.query(Bid).filter(Bid.team_id == team.id, Bid.ps_id == old.id, Bid.amount == 300).count() == 1
    assert db.get(WalletTransaction, transaction.id).amount == -300
    assert db.get(Team, team.id).coins == 700
    control = db.get(RoundControl, control.id)
    assert (control.round1_winning_bid_sum, control.round1_winning_bid_count) == (900, 3)


def test_duplicate_change_request_is_idempotent(client, admin_headers, db):
    old, target = _problem(db, 1), _problem(db, 2)
    team = _team(db, "Double Click Team", old)
    _round_control(db)
    db.commit()

    first = _change(client, admin_headers, team, target)
    second = _change(client, admin_headers, team, target)

    assert first.status_code == 200 and first.json()["idempotent"] is False
    assert second.status_code == 200 and second.json()["idempotent"] is True
    assert db.query(Team).filter(Team.id == team.id, Team.round1_problem_id == target.id).count() == 1
    assert db.query(WalletTransaction).filter(WalletTransaction.team_id == team.id).count() == 0


def test_change_remains_available_after_round_one_completion(client, admin_headers, db):
    old, target = _problem(db, 1), _problem(db, 5)
    team = _team(db, "Post Round Team", old)
    control = _round_control(db, ended=True)
    db.commit()

    response = _change(client, admin_headers, team, target)

    assert response.status_code == 200, response.text
    db.expire_all()
    assert db.get(Team, team.id).round1_problem_id == target.id
    control = db.get(RoundControl, control.id)
    assert control.ended is True and control.status == "CLOSED"


def test_two_admin_tabs_cannot_overfill_the_final_slot(tmp_path):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'change-problem-race.db'}",
        connect_args={"check_same_thread": False, "timeout": 10},
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    with factory() as db:
        target = _problem(db, 9)
        for index in range(4):
            _team(db, f"Occupied {index}", target)
        first = _team(db, "Final Slot A")
        second = _team(db, "Final Slot B")
        _round_control(db)
        db.commit()
        target_id, first_id, second_id = target.id, first.id, second.id

    def assign(team_id: int) -> str:
        with factory() as db:
            try:
                change_round1_problem_assignment(db, team_id, target_id)
                return "assigned"
            except Round1AssignmentError as exc:
                return str(exc)

    with ThreadPoolExecutor(max_workers=2) as executor:
        outcomes = list(executor.map(assign, (first_id, second_id)))

    assert outcomes.count("assigned") == 1
    assert sum("full (5/5)" in outcome for outcome in outcomes) == 1
    with factory() as db:
        assert db.query(Team).filter(Team.round1_problem_id == target_id).count() == 5
    engine.dispose()
