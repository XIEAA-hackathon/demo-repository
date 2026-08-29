import csv
import io

from openpyxl import load_workbook

from app.models.models import EventConfig, RoundControl, Team


def _twelve_team_registration() -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(["Team Name", "Leader Name", "Leader Email", "Leader Password", "Organizer Notes"])
    for index, letter in enumerate("ABCDEFGHIJKL", start=1):
        writer.writerow([f"Team {letter}", f"Leader {letter}", f"leader{index}@example.com", f"Leader{index}@123", f"Preserve note {letter}"])
    return output.getvalue().encode("utf-8")


def _credentials(client, admin_headers, content: bytes) -> dict[str, str]:
    imported = client.post(
        "/admin/registration/import",
        headers=admin_headers,
        files={"file": ("registrations.csv", content, "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    downloaded = client.get(
        f"/admin/registration/import/download/{imported.json()['download_token']}",
        headers=admin_headers,
    )
    assert downloaded.status_code == 200, downloaded.text
    return {
        row["Leader Login Email"]: row["Leader Password"]
        for row in csv.DictReader(io.StringIO(downloaded.content.decode("utf-8-sig")))
    }


def _login(client, credentials: dict[str, str], index: int) -> dict[str, str]:
    email = f"leader{index}@example.com"
    response = client.post("/login", data={"username": email, "password": credentials[email]})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _round_problem_csv() -> bytes:
    return (
        "Problem Number,Title,Description\n"
        "1,Round One Problem,First round description\n"
        "2,Round Two Problem,Second round description\n"
        "3,Final Round Problem,Final round description\n"
    ).encode()


def _assignment_rows(response) -> list[dict[str, str]]:
    assert response.status_code == 200, response.text
    return list(csv.DictReader(io.StringIO(response.content.decode("utf-8-sig"))))


def _assignment_workbook_rows(response) -> list[dict[str, object]]:
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    try:
        sheet = workbook["Participant Assignments"]
        values = list(sheet.iter_rows(values_only=True))
        headers = [str(value) for value in values[0]]
        return [dict(zip(headers, row)) for row in values[1:]]
    finally:
        workbook.close()


def test_top_five_lockout_base_prices_and_current_assignment_export(
    client, admin_headers, display_headers, db,
):
    credentials = _credentials(client, admin_headers, _twelve_team_registration())
    headers_by_index = {index: _login(client, credentials, index) for index in range(1, 13)}
    configured = client.put(
        "/admin/config",
        headers=admin_headers,
        json={"round1_minimum_bid": 150, "round1_bid_increment": 10, "bid_cooldown_seconds": 0},
    )
    assert configured.status_code == 200, configured.text

    imported = client.post(
        "/admin/rounds/round-1/problems/import",
        headers=admin_headers,
        files={"file": ("round1.csv", _round_problem_csv(), "text/csv")},
    )
    assert imported.status_code == 200, imported.text
    first_problem = next(problem for problem in imported.json()["problems"] if problem["problem_number"] == "1")
    second_problem = next(problem for problem in imported.json()["problems"] if problem["problem_number"] == "2")
    final_problem = next(problem for problem in imported.json()["problems"] if problem["problem_number"] == "3")

    assert client.post(f"/admin/rounds/round-1/problems/{first_problem['id']}/select", headers=admin_headers).status_code == 200
    assert client.post("/admin/rounds/round-1/preview/start", headers=admin_headers).status_code == 200
    assert client.post("/admin/rounds/round-1/bidding/start", headers=admin_headers).status_code == 200
    dashboard = client.get("/participant/dashboard", headers=headers_by_index[1]).json()
    assert dashboard["gameConfig"]["round1_minimum_bid"] == 150
    assert client.post("/bid", headers=headers_by_index[1], json={"ps_id": first_problem["id"], "amount": 149}).status_code == 422
    for index in (6, 5, 4, 3, 2, 1):
        response = client.post("/bid", headers=headers_by_index[index], json={"ps_id": first_problem["id"], "increment": 5})
        assert response.status_code == 200, response.text
    public_round = client.get("/public/leaderboard", headers=display_headers).json()
    assert public_round["base_price"] == 150
    assert client.post("/admin/rounds/round-1/bidding/close", headers=admin_headers).status_code == 200
    assigned = client.post("/admin/rounds/round-1/assign-winners", headers=admin_headers)
    assert assigned.status_code == 200, assigned.text
    assert [winner["team_name"] for winner in assigned.json()["winners"]] == [f"Team {letter}" for letter in "ABCDE"]
    db.expire_all()
    control = db.query(RoundControl).filter(RoundControl.round_type == "ROUND1").one()
    assert (control.round1_winning_bid_sum, control.round1_winning_bid_count) == (850, 5)
    duplicate = client.post("/admin/rounds/round-1/assign-winners", headers=admin_headers)
    assert duplicate.status_code == 200 and duplicate.json()["winners"] == []
    db.expire_all()
    control = db.query(RoundControl).filter(RoundControl.round_type == "ROUND1").one()
    assert (control.round1_winning_bid_sum, control.round1_winning_bid_count) == (850, 5)
    assert db.query(Team).filter(Team.team_name == "Team F").one().round1_problem_id is None

    after_first = _assignment_rows(client.get("/admin/registration/assignments", headers=admin_headers))
    first_by_team = {row["Team Name"]: row for row in after_first}
    assert first_by_team["Team A"]["Organizer Notes"] == "Preserve note A"
    assert first_by_team["Team A"]["Round 1 Problem Number"] == "R1-1"
    assert first_by_team["Team A"]["Round 1 Assignment Type"] == "BID_WINNER"
    assert first_by_team["Team A"]["Wildcard Problem Number"] == ""
    assert first_by_team["Team A"]["Final Problem Number"] == "R1-1"
    assert first_by_team["Team F"]["Round 1 Problem Number"] == ""
    assert first_by_team["Team A"]["Leader Password"] == "NOT EXPORTED"
    round_one_workbook = client.get("/admin/rounds/round-1/assignments/export", headers=admin_headers)
    assert 'filename="bid_to_build_round_1_assignments.xlsx"' in round_one_workbook.headers["content-disposition"]
    round_one_rows = _assignment_workbook_rows(round_one_workbook)
    round_one_by_team = {row["Team Name"]: row for row in round_one_rows}
    assert round_one_by_team["Team A"]["Round 1 Problem Title"] == "Round One Problem"
    assert round_one_by_team["Team F"]["Round 1 Problem Number"] in (None, "")

    updated = client.put("/admin/config", headers=admin_headers, json={"round1_minimum_bid": 200})
    assert updated.status_code == 200, updated.text
    assert client.post(f"/admin/rounds/round-1/problems/{second_problem['id']}/select", headers=admin_headers).status_code == 200
    assert client.post("/admin/rounds/round-1/preview/start", headers=admin_headers).status_code == 200
    assert client.post("/admin/rounds/round-1/bidding/start", headers=admin_headers).status_code == 200
    locked = client.post("/bid", headers=headers_by_index[1], json={"ps_id": second_problem["id"], "increment": 25})
    assert locked.status_code == 409
    assert "already has a Round 1 problem" in locked.json()["detail"]
    assert client.post("/bid", headers=headers_by_index[6], json={"ps_id": second_problem["id"], "amount": 199}).status_code == 422
    for index in (10, 9, 8, 7, 6):
        response = client.post("/bid", headers=headers_by_index[index], json={"ps_id": second_problem["id"], "increment": 5})
        assert response.status_code == 200, response.text
    assert client.post("/admin/rounds/round-1/bidding/close", headers=admin_headers).status_code == 200
    second_assigned = client.post("/admin/rounds/round-1/assign-winners", headers=admin_headers)
    assert [winner["team_name"] for winner in second_assigned.json()["winners"]] == [f"Team {letter}" for letter in "FGHIJ"]
    remaining = second_assigned.json()["remaining_problems"]
    assert remaining["suggested_deduction"] == 193
    final_row = next(row for row in remaining["problems"] if row["id"] == final_problem["id"])
    assert final_row["assignment_status"] == "UNASSIGNED"
    assert final_row["can_rebid"] is True and final_row["can_assign"] is True
    db.expire_all()
    control = db.query(RoundControl).filter(RoundControl.round_type == "ROUND1").one()
    assert (control.round1_winning_bid_sum, control.round1_winning_bid_count) == (1925, 10)
    final_teams = db.query(Team).filter(Team.team_name.in_(["Team K", "Team L"])).all()
    selected_team_ids = [team.id for team in final_teams]
    confirmed = client.post(
        f"/admin/rounds/round-1/problems/{final_problem['id']}/assign",
        headers=admin_headers,
        json={"team_ids": selected_team_ids, "deduction": 190},
    )
    assert confirmed.status_code == 200, confirmed.text
    db.expire_all()
    final_teams = db.query(Team).filter(Team.team_name.in_(["Team K", "Team L"])).all()
    assert all(team.round1_problem_id is not None for team in final_teams)
    assert all(team.round1_assignment_type == "MANUAL_ASSIGNMENT" for team in final_teams)
    assert all(team.round1_assignment_cost == 190 for team in final_teams)
    balances_after_confirmation = {team.id: team.coins for team in final_teams}
    duplicate_confirmation = client.post(
        f"/admin/rounds/round-1/problems/{final_problem['id']}/assign",
        headers=admin_headers,
        json={"team_ids": selected_team_ids, "deduction": 500},
    )
    assert duplicate_confirmation.status_code == 200
    db.expire_all()
    assert {
        team.id: team.coins
        for team in db.query(Team).filter(Team.team_name.in_(["Team K", "Team L"])).all()
    } == balances_after_confirmation
    control = db.query(RoundControl).filter(RoundControl.round_type == "ROUND1").one()
    assert (control.round1_winning_bid_sum, control.round1_winning_bid_count) == (1925, 10)

    assert client.post("/admin/rounds/round-1/end", headers=admin_headers).status_code == 200
    assert client.post("/admin/rounds/wildcard/applications/open", headers=admin_headers).status_code == 200
    assert client.post("/wildcard/apply", headers=headers_by_index[8]).status_code == 200
    assert client.post("/admin/rounds/wildcard/applications/close", headers=admin_headers).status_code == 200
    wildcard_import = client.post(
        "/admin/rounds/wildcard/problems/import",
        headers=admin_headers,
        files={"file": ("wildcard.csv", b"Problem Number,Title,Description\n2,Emergency Network,Wildcard description\n", "text/csv")},
    )
    assert wildcard_import.status_code == 200, wildcard_import.text
    wildcard_problem = wildcard_import.json()["problems"][0]
    wildcard_config = client.put(
        "/admin/config",
        headers=admin_headers,
        json={"wildcard_starting_bid": 300, "wildcard_bid_increment": 10},
    )
    assert wildcard_config.status_code == 200, wildcard_config.text
    assert client.post("/admin/rounds/wildcard/slots", headers=admin_headers, json={"slots": 1}).status_code == 200
    assert client.post("/admin/rounds/wildcard/bidding/start", headers=admin_headers).status_code == 200
    wildcard_dashboard = client.get("/participant/dashboard", headers=headers_by_index[8]).json()
    assert wildcard_dashboard["gameConfig"]["wildcard_starting_bid"] == 300
    assert client.post("/wildcard/bid", headers=headers_by_index[8], json={"amount": 300}).status_code == 422
    assert client.post("/wildcard/bid", headers=headers_by_index[8], json={"increment": 5}).status_code == 200
    public_wildcard = client.get("/public/leaderboard", headers=display_headers).json()
    assert public_wildcard["base_price"] == 300
    assert client.post("/admin/rounds/wildcard/bidding/close", headers=admin_headers).status_code == 200
    selected = client.post(f"/wildcard/select/{wildcard_problem['id']}", headers=headers_by_index[8])
    assert selected.status_code == 200, selected.text

    final_rows = _assignment_rows(client.get("/admin/registration/assignments", headers=admin_headers))
    team_h = next(row for row in final_rows if row["Team Name"] == "Team H")
    assert team_h["Round 1 Problem Number"] == "R1-2"
    assert team_h["Round 1 Problem Title"] == "Round Two Problem"
    team_k = next(row for row in final_rows if row["Team Name"] == "Team K")
    assert team_k["Round 1 Problem Number"] == "R1-3"
    assert team_k["Round 1 Assignment Type"] == "MANUAL_ASSIGNMENT"
    assert team_h["Wildcard Problem Number"] == "WC-2"
    assert team_h["Wildcard Problem Title"] == "Emergency Network"
    assert team_h["Final Problem Number"] == "WC-2"
    assert team_h["Final Problem Description"] == "Wildcard description"
    wildcard_workbook = client.get("/admin/rounds/wildcard/assignments/export", headers=admin_headers)
    assert 'filename="bid_to_build_wildcard_assignments.xlsx"' in wildcard_workbook.headers["content-disposition"]
    wildcard_rows = _assignment_workbook_rows(wildcard_workbook)
    wildcard_by_team = {row["Team Name"]: row for row in wildcard_rows}
    assert wildcard_by_team["Team H"]["Round 1 Problem Number"] == "R1-2"
    assert wildcard_by_team["Team H"]["Wildcard Problem Number"] == "WC-2"
    assert wildcard_by_team["Team A"]["Wildcard Problem Number"] in (None, "")

    reset = client.post("/admin/event-data/reset", headers=admin_headers, json={"confirmation": "RESET EVENT"})
    assert reset.status_code == 200, reset.text
    db.expire_all()
    config = db.query(EventConfig).one()
    assert config.round1_minimum_bid == 25
    assert config.wildcard_starting_bid == 150
    control = db.query(RoundControl).filter(RoundControl.round_type == "ROUND1").one()
    assert (control.round1_winning_bid_sum, control.round1_winning_bid_count) == (0, 0)
    assert client.get("/admin/registration/assignments", headers=admin_headers).status_code == 200
