"""Parsing + validation for organizer-provided Google Form Excel/CSV registration exports.

The exact spreadsheet column names are NOT hardcoded. Headers are normalized and
matched against a configurable alias table so a typical Google Form export
("Team Name", "Leader Email ID", "Member 1", ...) is detected automatically.
"""
import csv
import io
import json
import re
import secrets
import string
import time
from typing import Any, Dict, List

import pandas as pd

from app.core.security import is_valid_password_hash

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# --- canonical field -> accepted header aliases (normalized: lower, alnum only) ---
_HEADER_ALIASES: Dict[str, List[str]] = {
    "team_name": [
        "teamname", "teamname", "team", "nameofteam", "teamnameatheteam",
        "whatisthenameofyourteam", "teamnameyourtaketeam",
    ],
    "leader_name": [
        "leadername", "teamleader", "nameofleader", "teammemberleader",
        "leader", "teamleadername", "nameofteamleader", "leaderfullname",
        "whatisthenameoftheteamleader", "teamleadernameth",
    ],
    "leader_email": [
        "leaderemail", "teamleaderemail", "emailofleader", "leaderemailid",
        "emailid", "email", "leaderemailaddress", "emailidofteamleader",
        "whatistheemailidofyourteamleader", "teammembersemail",
    ],
    "leader_password": [
        "leaderpassword", "leaderloginpassword", "temporarypassword",
    ],
    "leader_password_hash": [
        "leaderpasswordhash", "leaderhash", "leaderloginpasswordhash", "passwordhash",
    ],
}

_MEMBER_NAME_PATTERN = re.compile(r"^(?:member|teammate|participant)\s*(\d+)(?:\s*(?:name|fullname))?$")
_MEMBER_EMAIL_PATTERN = re.compile(r"^(?:member|teammate|participant)\s*(\d+)(?:\s*(?:email|emailid))$")
_MEMBER_PASSWORD_PATTERN = re.compile(r"^(?:member|teammate|participant)\s*(\d+)(?:\s*(?:password|loginpassword))$")
_MEMBER_PASSWORD_HASH_PATTERN = re.compile(
    r"^(?:member|teammate|participant)\s*(\d+)(?:\s*(?:passwordhash|loginpasswordhash|hash))$"
)

def _norm_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]", "", header.lower())

def _detect_columns(headers: List[str]) -> Dict[str, Any]:
    """Map spreadsheet headers to canonical fields.

    Returns {"team_name": col_idx, "leader_name": idx, "leader_email": idx,
             "leader_password": idx, "member_names": {1: idx, ...},
             "member_emails": {1: idx, ...}, "member_passwords": {1: idx, ...}}
    and an "errors" list for anything unresolved.
    """
    mapping: Dict[str, Any] = {
        "col_to_family": {},
        "member_names": {},
        "member_emails": {},
        "member_passwords": {},
        "member_password_hashes": {},
        "team_name": None,
        "leader_name": None,
        "leader_email": None,
        "leader_password": None,
        "leader_password_hash": None,
    }
    normalized = [_norm_header(h) for h in headers]
    used_families: Dict[str, int] = {}

    for idx, norm in enumerate(normalized):
        if not norm:
            continue
        matched = False
        for family, aliases in _HEADER_ALIASES.items():
            if norm in aliases:
                mapping[family] = idx
                mapping["col_to_family"][idx] = family
                used_families[family] = idx
                matched = True
                break
        if matched:
            continue

        m = _MEMBER_NAME_PATTERN.match(norm)
        if m:
            number = int(m.group(1))
            if number >= 1:
                if number not in mapping["member_names"]:
                    mapping["member_names"][number] = idx
                    mapping["col_to_family"][idx] = f"member_name:{number}"
                matched = True
                continue

        m = _MEMBER_EMAIL_PATTERN.match(norm)
        if m:
            number = int(m.group(1))
            if number >= 1:
                if number not in mapping["member_emails"]:
                    mapping["member_emails"][number] = idx
                    mapping["col_to_family"][idx] = f"member_email:{number}"
                matched = True
                continue

        m = _MEMBER_PASSWORD_PATTERN.match(norm)
        if m:
            number = int(m.group(1))
            if number >= 1:
                if number not in mapping["member_passwords"]:
                    mapping["member_passwords"][number] = idx
                    mapping["col_to_family"][idx] = f"member_password:{number}"
                continue

        m = _MEMBER_PASSWORD_HASH_PATTERN.match(norm)
        if m:
            number = int(m.group(1))
            if number >= 1 and number not in mapping["member_password_hashes"]:
                mapping["member_password_hashes"][number] = idx
                mapping["col_to_family"][idx] = f"member_password_hash:{number}"

    errors = []
    if mapping["team_name"] is None:
        errors.append("Could not detect a 'Team Name' column.")
    if mapping["leader_name"] is None:
        errors.append("Could not detect a 'Leader Name' column.")
    if mapping["leader_email"] is None:
        errors.append("Could not detect a 'Leader Email' column.")
    return mapping, errors

def _is_valid_email(value: str) -> bool:
    return bool(EMAIL_RE.match(value.strip()))

def _default_password() -> str:
    """Return a one-time password with mixed character classes.

    The plaintext value is returned to the administrator once; callers persist
    only its password hash.
    """
    required = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
        secrets.choice("!@#$%"),
    ]
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    password = required + [secrets.choice(alphabet) for _ in range(10)]
    secrets.SystemRandom().shuffle(password)
    return "".join(password)

def _read_registration_dataframe(filename: str, content: bytes) -> pd.DataFrame:
    """Read an organizer registration file into the parser's core representation."""
    source = io.BytesIO(content)
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return pd.read_csv(
            source,
            dtype=str,
            index_col=False,
            encoding="utf-8-sig",
            encoding_errors="replace",
        )
    if lower_name.endswith((".xlsx", ".xlsm")):
        return pd.read_excel(source, dtype=str, engine="openpyxl")
    raise ValueError("Unsupported file type. Upload .xlsx or .csv.")


def _column_series(frame: pd.DataFrame, mapping: Dict[str, Any], family: str) -> pd.Series:
    index = mapping.get(family)
    if index is None:
        return pd.Series("", index=frame.index, dtype="object")
    return frame.iloc[:, index]


def parse_registration_file(filename: str, content: bytes, simple_mode: bool = False) -> Dict[str, Any]:
    """Parse and validate an uploaded registration workbook without mutating data."""
    del simple_mode  # Retained for the existing caller contract.
    empty_result = {
        "rows": [], "warnings": [], "errors": [], "row_errors": [],
        "detected_columns": {}, "source_headers": [],
        "timings": {"pandas_parse": 0.0, "column_normalization": 0.0, "validation": 0.0},
        "credential_inputs": {"supplied_hashes": 0, "plaintext_passwords": 0, "without_credentials": 0},
    }

    parse_started_at = time.perf_counter()
    try:
        frame = _read_registration_dataframe(filename, content)
    except ValueError as exc:
        return {**empty_result, "errors": [str(exc)]}
    except Exception as exc:  # pragma: no cover - defensive
        return {**empty_result, "errors": [f"Could not read file: {exc}"]}
    pandas_parse_duration = time.perf_counter() - parse_started_at

    normalization_started_at = time.perf_counter()
    headers = [str(column).strip() for column in frame.columns]
    normalized_to_original = {_norm_header(column): column for column in headers}
    mapping, header_errors = _detect_columns(headers)
    text_frame = frame.fillna("").astype(str)
    for column_index in range(len(text_frame.columns)):
        text_frame.iloc[:, column_index] = text_frame.iloc[:, column_index].str.strip()
    column_normalization_duration = time.perf_counter() - normalization_started_at
    timings = {
        "pandas_parse": pandas_parse_duration,
        "column_normalization": column_normalization_duration,
        "validation": 0.0,
    }

    if text_frame.empty or not text_frame.ne("").any(axis=None):
        return {**empty_result, "errors": ["The file contains no rows."], "timings": timings}
    if header_errors:
        return {
            **empty_result,
            "errors": list(header_errors),
            "detected_columns": {**mapping, "normalized_to_original": normalized_to_original},
            "source_headers": headers,
            "timings": timings,
        }

    validation_started_at = time.perf_counter()
    team_names = _column_series(text_frame, mapping, "team_name")
    team_keys = team_names.str.lower()
    leader_names = _column_series(text_frame, mapping, "leader_name")
    leader_emails = _column_series(text_frame, mapping, "leader_email").str.lower()
    leader_passwords = _column_series(text_frame, mapping, "leader_password")
    leader_password_hashes = _column_series(text_frame, mapping, "leader_password_hash")
    nonempty_rows = text_frame.ne("").any(axis=1)
    missing_team_names = team_names.eq("")
    missing_leader_names = leader_names.eq("")
    missing_leader_emails = leader_emails.eq("")
    invalid_leader_emails = leader_emails.ne("") & ~leader_emails.str.match(EMAIL_RE)
    duplicate_teams = team_keys.ne("") & team_keys.duplicated(keep="first")
    duplicate_leaders = leader_emails.ne("") & leader_emails.duplicated(keep="first")

    first_team_rows: Dict[str, int] = {}
    first_leader_rows: Dict[str, int] = {}
    for position in range(len(text_frame.index)):
        spreadsheet_row = position + 2
        if team_keys.iat[position]:
            first_team_rows.setdefault(team_keys.iat[position], spreadsheet_row)
        if leader_emails.iat[position]:
            first_leader_rows.setdefault(leader_emails.iat[position], spreadsheet_row)

    member_numbers = sorted(
        set(mapping["member_names"])
        | set(mapping["member_emails"])
        | set(mapping["member_passwords"])
        | set(mapping["member_password_hashes"])
    )
    member_names = {
        number: text_frame.iloc[:, mapping["member_names"][number]]
        if number in mapping["member_names"] else pd.Series("", index=text_frame.index, dtype="object")
        for number in member_numbers
    }
    member_emails = {
        number: (
            text_frame.iloc[:, mapping["member_emails"][number]].str.lower()
            if number in mapping["member_emails"] else pd.Series("", index=text_frame.index, dtype="object")
        )
        for number in member_numbers
    }
    member_passwords = {
        number: text_frame.iloc[:, mapping["member_passwords"][number]]
        if number in mapping["member_passwords"] else pd.Series("", index=text_frame.index, dtype="object")
        for number in member_numbers
    }
    member_password_hashes = {
        number: text_frame.iloc[:, mapping["member_password_hashes"][number]]
        if number in mapping["member_password_hashes"] else pd.Series("", index=text_frame.index, dtype="object")
        for number in member_numbers
    }
    leader_hash_mask = leader_password_hashes.ne("")
    member_hash_masks = {
        number: member_password_hashes[number].ne("") for number in member_numbers
    }
    sensitive_column_indexes = {
        index for index, family in mapping["col_to_family"].items()
        if family in {"leader_password", "leader_password_hash"}
        or family.startswith("member_password:")
        or family.startswith("member_password_hash:")
    }

    errors: List[str] = []
    warnings: List[str] = []
    row_errors: List[Dict[str, Any]] = []
    rows: List[Dict[str, Any]] = []
    seen_identities: Dict[str, Dict[str, Any]] = {}
    credential_inputs = {"supplied_hashes": 0, "plaintext_passwords": 0, "without_credentials": 0}

    for position in range(len(text_frame.index)):
        if not nonempty_rows.iat[position]:
            continue
        row_number = position + 2
        cells = text_frame.iloc[position].tolist()
        team_name = team_names.iat[position]
        leader_name = leader_names.iat[position]
        leader_email = leader_emails.iat[position]
        leader_password = leader_passwords.iat[position]
        leader_password_hash = leader_password_hashes.iat[position]
        messages: List[str] = []

        missing = [
            label for label, is_missing in (
                ("team name", missing_team_names.iat[position]),
                ("leader name", missing_leader_names.iat[position]),
                ("leader email", missing_leader_emails.iat[position]),
            ) if is_missing
        ]
        if missing:
            messages.append(f"Missing required field(s): {', '.join(missing)}.")
        if invalid_leader_emails.iat[position]:
            messages.append(f"Leader email '{leader_email}' is not a valid email.")
        if duplicate_teams.iat[position]:
            messages.append(
                f"Team '{team_name}' already appears at row {first_team_rows[team_keys.iat[position]]}."
            )
        if duplicate_leaders.iat[position]:
            messages.append(
                f"Leader email '{leader_email}' already appears at row {first_leader_rows[leader_email]}."
            )
        if leader_hash_mask.iat[position] and not is_valid_password_hash(leader_password_hash):
            messages.append("Leader Password Hash is not a structurally valid sha256$salt$digest hash.")

        members: List[Dict[str, Any]] = []
        for number in member_numbers:
            member_name = member_names[number].iat[position]
            member_email = member_emails[number].iat[position]
            member_password = member_passwords[number].iat[position]
            member_password_hash = member_password_hashes[number].iat[position]
            if not member_name and not member_email and not member_password and not member_password_hash:
                continue
            if member_email and not member_name:
                messages.append(f"Member {number} has an email but no name.")
            if member_password and not member_email:
                messages.append(f"Member {number} has a password but no email/login ID.")
            if member_password_hash and not member_email:
                messages.append(f"Member {number} has a password hash but no email/login ID.")
            if member_email and not _is_valid_email(member_email):
                messages.append(f"Member {number} email '{member_email}' is not a valid email.")
            if member_hash_masks[number].iat[position] and not is_valid_password_hash(member_password_hash):
                messages.append(
                    f"Member {number} Password Hash is not a structurally valid sha256$salt$digest hash."
                )
            members.append({
                "number": number,
                "name": member_name,
                "email": member_email,
                "password": member_password,
                "password_hash": member_password_hash,
            })

        identities = [(leader_email, "leader")]
        identities.extend((member["email"], "member") for member in members if member["email"])
        row_identity_emails: set[str] = set()
        for identity_email, identity_role in identities:
            if not identity_email:
                continue
            if identity_email in row_identity_emails:
                messages.append(f"Email '{identity_email}' is reused within this team row.")
                continue
            row_identity_emails.add(identity_email)
            previous = seen_identities.get(identity_email)
            if previous and previous["row_number"] != row_number:
                messages.append(
                    f"Email '{identity_email}' is already used as {previous['role']} for "
                    f"team '{previous['team_name']}' at row {previous['row_number']}."
                )
            else:
                seen_identities[identity_email] = {
                    "row_number": row_number,
                    "team_name": team_name,
                    "role": identity_role,
                }

        if messages:
            for message in messages:
                errors.append(f"Row {row_number}: {message}")
                row_errors.append({"row_number": row_number, "message": message})
            continue

        participants = [{"password": leader_password, "password_hash": leader_password_hash}, *members]
        for participant in participants:
            if participant.get("password_hash"):
                credential_inputs["supplied_hashes"] += 1
            elif participant.get("password"):
                credential_inputs["plaintext_passwords"] += 1
            else:
                credential_inputs["without_credentials"] += 1

        rows.append({
            "row_number": row_number,
            "team_name": team_name,
            "leader_name": leader_name,
            "leader_email": leader_email,
            "leader_password": leader_password,
            "leader_password_hash": leader_password_hash,
            "members": members,
            "warnings": [],
            "status": "new",
            "source_values": [
                "EXISTING ACCOUNT" if index in sensitive_column_indexes and value else value
                for index, value in enumerate(cells)
            ],
        })

    timings["validation"] = time.perf_counter() - validation_started_at
    return {
        "rows": rows,
        "warnings": warnings,
        "errors": errors,
        "row_errors": row_errors,
        "detected_columns": {**mapping, "normalized_to_original": normalized_to_original},
        "source_headers": headers,
        "timings": timings,
        "credential_inputs": credential_inputs,
    }


def build_registration_credential_workbook(
    filename: str,
    content: bytes,
    leader_credentials: Dict[int, Dict[str, str]],
) -> bytes:
    """Preserve the uploaded sheet and append login plus credential status."""
    from copy import copy
    from io import BytesIO
    from openpyxl import Workbook, load_workbook

    if filename.lower().endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(BytesIO(content))
        sheet = workbook.active
        hash_columns = [
            column
            for column in range(1, sheet.max_column + 1)
            if _is_password_hash_header(sheet.cell(row=1, column=column).value)
        ]
        for row_number in range(2, sheet.max_row + 1):
            for column in hash_columns:
                if sheet.cell(row=row_number, column=column).value:
                    sheet.cell(row=row_number, column=column, value="NOT EXPORTED")
    else:
        workbook = Workbook()
        sheet = workbook.active
        frame = _credential_export_frame(filename, content)
        sheet.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            sheet.append(list(row))

    login_column = sheet.max_column + 1
    status_column = login_column + 1
    sheet.cell(row=1, column=login_column, value="Leader Login Email")
    sheet.cell(row=1, column=status_column, value="Credential Status")

    if login_column > 1:
        source_header = sheet.cell(row=1, column=login_column - 1)
        for column in (login_column, status_column):
            target = sheet.cell(row=1, column=column)
            target._style = copy(source_header._style)
            target.font = copy(source_header.font)
            target.fill = copy(source_header.fill)
            target.border = copy(source_header.border)
            target.alignment = copy(source_header.alignment)
            target.number_format = source_header.number_format

    for row_number, credential in leader_credentials.items():
        sheet.cell(row=row_number, column=login_column, value=credential["email"])
        sheet.cell(row=row_number, column=status_column, value=credential["status"])

    sheet.column_dimensions[sheet.cell(row=1, column=login_column).column_letter].width = 34
    sheet.column_dimensions[sheet.cell(row=1, column=status_column).column_letter].width = 24
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def build_registration_credential_csv(
    content: bytes,
    leader_credentials: Dict[int, Dict[str, str]],
) -> bytes:
    """Preserve CSV cell values and append login plus credential status."""
    frame = _credential_export_frame("registration.csv", content)
    if frame.empty:
        return b""
    frame["Leader Login Email"] = ""
    frame["Credential Status"] = ""
    for position in range(len(frame.index)):
        row_number = position + 2
        credential = leader_credentials.get(row_number, {"email": "", "status": ""})
        frame.iat[position, frame.columns.get_loc("Leader Login Email")] = credential["email"]
        frame.iat[position, frame.columns.get_loc("Credential Status")] = credential["status"]
    return frame.to_csv(index=False, lineterminator="\r\n").encode("utf-8-sig")


def _is_password_hash_header(value: Any) -> bool:
    normalized = _norm_header(str(value or ""))
    return normalized in {"leaderhash", "passwordhash"} or normalized.endswith("passwordhash")


def _credential_export_frame(filename: str, content: bytes) -> pd.DataFrame:
    frame = _read_registration_dataframe(filename, content).fillna("")
    for column in frame.columns:
        if _is_password_hash_header(column):
            populated = frame[column].astype(str).str.strip().ne("")
            frame.loc[populated, column] = "NOT EXPORTED"
    return frame


ASSIGNMENT_HEADERS = [
    "Round 1 Assigned Problem",
    "Wildcard Assigned Problem",
    "GitHub Link",
    "Round 1 Problem Number",
    "Round 1 Problem Title",
    "Round 1 Problem Description",
    "Round 1 Assignment Type",
    "Round 1 Final Price / Winning Bid",
    "Wildcard Problem Number",
    "Wildcard Problem Title",
    "Wildcard Problem Description",
    "Wildcard Final Price / Winning Bid",
    "Final Problem Number",
    "Final Problem Title",
    "Final Problem Description",
]


def _spreadsheet_safe(value: Any) -> Any:
    """Keep user-controlled text from being interpreted as a spreadsheet formula."""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def build_registration_assignment_csv(headers: List[str], rows: List[List[Any]]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow([_spreadsheet_safe(value) for value in headers])
    writer.writerows([[_spreadsheet_safe(value) for value in row] for row in rows])
    return output.getvalue().encode("utf-8-sig")


def build_registration_assignment_workbook(headers: List[str], rows: List[List[Any]]) -> bytes:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Participant Assignments"
    sheet.append([_spreadsheet_safe(value) for value in headers])
    for row in rows:
        sheet.append([_spreadsheet_safe(value) for value in row])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column]
        sheet.column_dimensions[column[0].column_letter].width = min(60, max(12, max(map(len, values), default=0) + 2))
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
