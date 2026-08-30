from __future__ import annotations

import csv
import io
import logging
import re
from typing import Iterable

from fastapi import APIRouter, Depends, File, HTTPException, Response as FastAPIResponse, UploadFile
from fastapi.responses import Response
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.orm import Session

from app.api.auth import get_current_active_admin, get_current_active_display, get_current_user
from app.api.websockets import manager
from app.core.database import get_db
from app.models.models import Bid, GameConfig, ProblemStatement, RoundControl, Team, WalletTransaction, Wildcard
from app.services.event_service import (
    _remaining_seconds,
    event_snapshot,
    get_or_create_event_config,
    get_or_create_game_config,
    get_or_create_round_control,
    sync_expired_event_state,
    transition_event_state,
)
from app.services.activity_log import record_event
from app.core.event_constants import ROUND1_WINNER_COUNT
from app.services.round1_assignment import (
    EXTERNAL_PROBLEM_ROUND,
    ROUND1_BID_WINNER,
    ROUND1_FINALIZATION_LOCK,
    ROUND1_PROBLEM_CAPACITY,
    Round1AssignmentError,
    change_round1_problem_assignment,
    manually_assign_problem,
    round1_assignment_management_payload,
    remaining_problems_payload,
    update_round1_winning_bid_aggregate,
)
from app.services.wildcard_service import ranking_payload, wildcard_payload

router = APIRouter()
logger = logging.getLogger("uvicorn.error")

ROUND_META = {
    "round-1": {"type": "ROUND1", "number": 1, "prefix": "R1", "label": "Round 1"},
    "wildcard": {"type": "WILDCARD", "number": 2, "prefix": "WC", "label": "Wildcard"},
}
NUMBER_HEADERS = {"problemnumber", "problemno", "problemid", "number", "id"}
TITLE_HEADERS = {"title", "problemtitle", "name", "problemname"}
DESCRIPTION_HEADERS = {"description", "problemdescription", "problemstatement", "statement", "problem"}


class ManualProblemAssignmentRequest(BaseModel):
    team_ids: list[int] = Field(min_length=1)
    deduction: int = Field(ge=0, strict=True)


class ChangeProblemAssignmentRequest(BaseModel):
    target_problem_id: int = Field(gt=0, strict=True)
    new_balance: int | None = Field(default=None, ge=0, le=1_000_000, strict=True)


def _meta(round_slug: str) -> dict:
    meta = ROUND_META.get(round_slug)
    if not meta:
        raise HTTPException(status_code=404, detail="Round not found")
    return meta


def _normalized(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _read_problem_rows(filename: str, content: bytes) -> list[tuple[int, str, str]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="Unsupported file. Upload an .xlsx or .csv file.")
    if not content:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")

    if suffix == "csv":
        try:
            rows: Iterable[list[object]] = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
        except (UnicodeDecodeError, csv.Error) as exc:
            raise HTTPException(status_code=400, detail="The CSV file could not be read.") from exc
    else:
        try:
            sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
            rows = list(sheet.iter_rows(values_only=True))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="The XLSX file could not be read.") from exc

    rows = [list(row) for row in rows if any(str(cell or "").strip() for cell in row)]
    if not rows:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    headers = [_normalized(value) for value in rows[0]]
    number_index = next((index for index, value in enumerate(headers) if value in NUMBER_HEADERS), None)
    title_index = next((index for index, value in enumerate(headers) if value in TITLE_HEADERS), None)
    description_index = next((index for index, value in enumerate(headers) if value in DESCRIPTION_HEADERS), None)
    if number_index is None:
        raise HTTPException(status_code=400, detail="Missing problem number column.")
    if title_index is None:
        raise HTTPException(status_code=400, detail="Missing problem title column.")
    if description_index is None:
        raise HTTPException(status_code=400, detail="Missing problem description column.")

    parsed: list[tuple[int, str, str]] = []
    seen: set[int] = set()
    errors: list[str] = []
    for source_row, row in enumerate(rows[1:], start=2):
        raw_number = row[number_index] if number_index < len(row) else None
        title = str(row[title_index] if title_index < len(row) else "").strip()
        description = str(row[description_index] if description_index < len(row) else "").strip()
        try:
            number = int(raw_number)
            if number <= 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"Row {source_row}: problem number must be a positive whole number.")
            continue
        if number in seen:
            errors.append(f"Row {source_row}: duplicate problem number {number}.")
        if not title:
            errors.append(f"Row {source_row}: problem title is required.")
        if not description:
            errors.append(f"Row {source_row}: problem description is required.")
        if number not in seen and title and description:
            parsed.append((number, title, description))
        seen.add(number)
    if not parsed and not errors:
        errors.append("The file contains headers but no problems.")
    if errors:
        raise HTTPException(status_code=422, detail=errors)
    return parsed


def _display_number(problem: ProblemStatement) -> str:
    return problem.ps_number.split("-", 1)[-1]


def _problem_payload(problem: ProblemStatement, control: RoundControl) -> dict:
    if problem.id == control.current_problem_id:
        status = "CURRENT"
    elif problem.status in {"completed", "allocated"}:
        status = "COMPLETED"
    elif problem.status == "no_bids":
        status = "NO_BIDS"
    else:
        status = "AVAILABLE"
    return {
        "id": problem.id,
        "problem_number": _display_number(problem),
        "title": problem.title,
        "description": problem.description,
        "problem_statement": problem.description or problem.title,
        "status": status,
    }


def _sync_application_window(db: Session, control: RoundControl) -> None:
    if control.applications_open and _remaining_seconds(get_or_create_game_config(db)) == 0:
        control.applications_open = False
        db.commit()


def _round_payload(db: Session, meta: dict) -> dict:
    if meta["type"] == "WILDCARD":
        return wildcard_payload(db)
    control = get_or_create_round_control(db, meta["type"])
    config = get_or_create_event_config(db)
    problems = db.query(ProblemStatement).filter(ProblemStatement.round == meta["number"]).order_by(ProblemStatement.id).all()
    current = next((problem for problem in problems if problem.id == control.current_problem_id), None)
    current_bids = []
    if current:
        current_bids = db.query(Bid).filter(Bid.ps_id == current.id, Bid.round == meta["number"]).order_by(Bid.amount.desc(), Bid.timestamp.asc()).all()
    highest = current_bids[0] if current_bids else None
    highest_team = db.query(Team).filter(Team.id == highest.team_id).first() if highest else None
    remaining_problems = remaining_problems_payload(db, control)
    payload = {
        "round_type": meta["type"],
        "status": control.status,
        "ended": control.ended,
        "current_problem": _problem_payload(current, control) if current else None,
        "problems": [_problem_payload(problem, control) for problem in problems],
        "highest_bid": highest.amount if highest else None,
        "highest_team": highest_team.team_name if highest_team else None,
        "settings": {
            "preview_seconds": config.round1_preview_seconds if meta["number"] == 1 else config.wildcard_preview_seconds,
            "bidding_seconds": config.round1_bid_seconds if meta["number"] == 1 else config.wildcard_bid_seconds,
            "base_price": config.round1_minimum_bid,
            "winner_count": ROUND1_WINNER_COUNT,
        },
        "remaining_problems": remaining_problems,
        "assigned_team_count": db.query(Team).filter(Team.round1_problem_id.is_not(None)).count(),
        "unassigned_team_count": db.query(Team).filter(
            Team.is_approved.is_(True), Team.is_system_team.is_(False), Team.round1_problem_id.is_(None)
        ).count(),
        "event": event_snapshot(db),
    }
    return payload


@router.get("/admin/rounds/{round_slug}")
def get_round(round_slug: str, db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    return _round_payload(db, _meta(round_slug))


@router.post("/admin/rounds/{round_slug}/problems/import")
async def import_problems(round_slug: str, file: UploadFile = File(...), db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    meta = _meta(round_slug)
    rows = _read_problem_rows(file.filename or "", await file.read())
    created = updated = 0
    for number, title, description in rows:
        internal_number = f"{meta['prefix']}-{number}"
        problem = db.query(ProblemStatement).filter(ProblemStatement.ps_number == internal_number).first()
        if problem and problem.status in {"current", "completed", "allocated"}:
            raise HTTPException(status_code=409, detail=f"Problem {number} is already current or completed and cannot be overwritten.")
        if problem:
            problem.title = title
            problem.description = description
            problem.status = "available"
            updated += 1
        else:
            db.add(ProblemStatement(ps_number=internal_number, title=title, description=description, round=meta["number"], status="available"))
            created += 1
    control = get_or_create_round_control(db, meta["type"])
    if meta["type"] == "ROUND1" and control.status == "IDLE":
        control.status = "READY"
    record_event(db, f"{meta['type'].lower()}.problems_imported", actor=current_user, metadata={"count": len(rows), "filename": file.filename or ""})
    db.commit()
    response = {"imported": len(rows), "created": created, "updated": updated, **_round_payload(db, meta)}
    db.close()
    await manager.broadcast_event("round_updated", {"round": meta["type"], "action": "problems_imported"})
    return response


@router.get("/admin/rounds/{round_slug}/problems/sample.csv")
def problem_sample(round_slug: str, current_user=Depends(get_current_active_admin)):
    meta = _meta(round_slug)
    sample = (
        "Problem Number,Title,Description\r\n"
        "1,\"Adaptive Noise Cancellation\",\"Develop an AI/ML-enabled adaptive noise cancellation system...\"\r\n"
        "2,\"Tropical Cyclone Prediction\",\"Develop a system using multi-source satellite data...\"\r\n"
        "3,\"Emergency Communication\",\"Build a lightweight emergency communication solution...\"\r\n"
    )
    return Response(sample, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={round_slug}-problems-sample.csv"})


@router.post("/admin/rounds/{round_slug}/problems/{problem_id}/select")
async def select_problem(round_slug: str, problem_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    meta = _meta(round_slug)
    if meta["type"] == "WILDCARD":
        raise HTTPException(status_code=409, detail="Wildcard problems are selected by qualified teams after slot bidding.")
    control = get_or_create_round_control(db, meta["type"])
    if control.current_problem_id == problem_id:
        return _round_payload(db, meta)
    if control.ended:
        raise HTTPException(status_code=409, detail=f"{meta['label']} is closed.")
    if control.current_problem_id:
        raise HTTPException(status_code=409, detail="Complete the current problem before selecting another one.")
    problem = db.query(ProblemStatement).filter(ProblemStatement.id == problem_id, ProblemStatement.round == meta["number"]).first()
    if not problem:
        raise HTTPException(status_code=404, detail="Problem not found in this round.")
    if problem.status not in {"available", "visible"}:
        raise HTTPException(status_code=409, detail="A completed problem cannot be selected again.")
    # A completed auction leaves the global event in ROUND1_RESULT. Reset that
    # transient state when the next independent auction is selected so clients
    # do not mistake the new READY problem for the previous result.
    transition_event_state(db, "WAITING", validate=False, commit=False)
    problem.status = "current"
    control.current_problem_id = problem.id
    control.status = "READY"
    record_event(db, "round1.problem_selected", actor=current_user, entity_type="problem", entity_id=problem.id)
    db.commit()
    problem_id_value = problem.id
    snapshot = event_snapshot(db)
    response = _round_payload(db, meta)
    db.close()
    await manager.broadcast_event("round_updated", {"round": meta["type"], "action": "problem_selected", "problem_id": problem_id_value})
    await manager.broadcast_event("event_state_changed", snapshot)
    return response


@router.post("/admin/rounds/{round_slug}/preview/start")
async def start_preview(round_slug: str, db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    meta = _meta(round_slug)
    if meta["type"] == "WILDCARD":
        raise HTTPException(status_code=409, detail="Wildcard no longer has a global problem preview phase.")
    control = get_or_create_round_control(db, meta["type"])
    if control.status in {"PREVIEW", "PREVIEW_EXPIRED"}:
        return _round_payload(db, meta)
    if control.ended or not control.current_problem_id:
        raise HTTPException(status_code=409, detail="Select an available problem before starting preview.")
    state = "ROUND1_PREVIEW" if meta["number"] == 1 else "WILDCARD_PREVIEW"
    transition_event_state(db, state, validate=False, commit=False)
    control.status = "PREVIEW"
    record_event(db, "round1.preview_started", actor=current_user, entity_type="problem", entity_id=control.current_problem_id)
    db.commit()
    snapshot = event_snapshot(db)
    response = _round_payload(db, meta)
    db.close()
    await manager.broadcast_event("event_state_changed", snapshot)
    return response


@router.post("/admin/rounds/{round_slug}/bidding/start")
async def start_bidding(round_slug: str, db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    meta = _meta(round_slug)
    if meta["type"] == "WILDCARD":
        raise HTTPException(status_code=409, detail="Use the Wildcard slot bidding control.")
    control = get_or_create_round_control(db, meta["type"])
    sync_expired_event_state(db)
    db.refresh(control)
    if control.status == "BIDDING":
        return _round_payload(db, meta)
    if control.ended or not control.current_problem_id or control.status not in {"PREVIEW", "PREVIEW_EXPIRED", "READY"}:
        raise HTTPException(status_code=409, detail="Preview a selected problem before starting bidding.")
    state = "ROUND1_BIDDING" if meta["number"] == 1 else "WILDCARD_BIDDING"
    transition_event_state(db, state, validate=False, commit=False)
    control.status = "BIDDING"
    record_event(db, "round1.bidding_started", actor=current_user, entity_type="problem", entity_id=control.current_problem_id)
    db.commit()
    snapshot = event_snapshot(db)
    response = _round_payload(db, meta)
    db.close()
    await manager.broadcast_event("event_state_changed", snapshot)
    return response


@router.post("/admin/rounds/{round_slug}/bidding/close")
async def close_bidding(round_slug: str, db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    meta = _meta(round_slug)
    if meta["type"] == "WILDCARD":
        raise HTTPException(status_code=409, detail="Use the Wildcard slot bidding control.")
    control = get_or_create_round_control(db, meta["type"])
    sync_expired_event_state(db)
    db.refresh(control)
    game = get_or_create_game_config(db)
    if control.status == "READY" and game.state == "ROUND1_RESULT":
        logger.info("Duplicate Round 1 close request ignored; bidding is already closed.")
        return _round_payload(db, meta)
    if control.status != "BIDDING":
        raise HTTPException(status_code=409, detail="Bidding is not active.")
    transition_event_state(db, "ROUND1_RESULT" if meta["number"] == 1 else "WILDCARD_SELECTION", validate=False, commit=False)
    control.status = "READY"
    record_event(db, "round1.bidding_closed", actor=current_user, entity_type="problem", entity_id=control.current_problem_id)
    db.commit()
    # Materialize every DB-backed value while the session is available, then
    # release its connection before potentially slow WebSocket fan-out.
    snapshot = event_snapshot(db)
    response = _round_payload(db, meta)
    db.close()
    await manager.broadcast_event("event_state_changed", snapshot)
    return response


@router.post("/admin/rounds/{round_slug}/assign-winners")
async def assign_winners(round_slug: str, db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    meta = _meta(round_slug)
    if meta["type"] == "WILDCARD":
        raise HTTPException(status_code=409, detail="Wildcard slot winners are finalized when slot bidding closes.")
    with ROUND1_FINALIZATION_LOCK:
        unlocked_control = get_or_create_round_control(db, meta["type"])
        control = (
            db.query(RoundControl)
            .filter(RoundControl.id == unlocked_control.id)
            .with_for_update()
            .populate_existing()
            .one()
        )
        if control.current_problem_id is None and control.status in {"READY", "COMPLETE"}:
            logger.info("Duplicate Round 1 winner assignment request ignored; assignment is already complete.")
            return {"message": "Winner assignment already completed.", "winners": [], **_round_payload(db, meta)}
        problem = db.query(ProblemStatement).filter(ProblemStatement.id == control.current_problem_id, ProblemStatement.round == meta["number"]).first()
        if not problem or control.status != "READY":
            raise HTTPException(status_code=409, detail="Close bidding before assigning winners.")
        event_config = get_or_create_event_config(db)
        existing_assignments = db.query(Team).filter(Team.round1_problem_id == problem.id).count()
        winner_limit = (
            max(0, ROUND1_PROBLEM_CAPACITY - existing_assignments)
            if meta["number"] == 1
            else event_config.wildcard_slots
        )
        bids = db.query(Bid).filter(Bid.ps_id == problem.id, Bid.round == meta["number"]).order_by(Bid.amount.desc(), Bid.timestamp.asc(), Bid.team_id.asc()).all()
        winners = []
        for bid in bids:
            if len(winners) >= winner_limit:
                break
            team = db.query(Team).filter(Team.id == bid.team_id).first()
            if not team or team.round1_problem_id is not None or team.ps_id is not None or team.coins < bid.amount:
                continue
            if meta["number"] == 2:
                application = db.query(Wildcard).filter(Wildcard.team_id == team.id, Wildcard.status == "applied").first()
                if not application:
                    continue
                application.status = "selected"
                application.used = True
                application.coins_paid = bid.amount
            team.coins -= bid.amount
            team.ps_id = problem.id
            if meta["number"] == 1:
                team.round1_problem_id = problem.id
                team.round1_assignment_type = ROUND1_BID_WINNER
                team.round1_assignment_cost = bid.amount
            db.add(WalletTransaction(team_id=team.id, transaction_type="ROUND1_WIN" if meta["number"] == 1 else "WILDCARD_WIN", amount=-bid.amount, description=f"{meta['label']} win for problem {_display_number(problem)}"))
            winners.append({"team_id": team.id, "team_name": team.team_name, "amount": bid.amount})
        if meta["number"] == 1:
            update_round1_winning_bid_aggregate(
                control,
                problem,
                [winner["amount"] for winner in winners],
            )
        problem.status = "completed" if winners else "no_bids"
        assigned_problem = {
            "id": problem.id,
            "number": int(_display_number(problem)) if _display_number(problem).isdigit() else problem.ps_number,
            "title": problem.title,
            "description": problem.description or "",
            "starting_bid": get_or_create_event_config(db).round1_minimum_bid,
        }
        control.current_problem_id = None
        unassigned_count = db.query(Team).filter(
            Team.is_approved.is_(True),
            Team.is_system_team.is_(False),
            Team.round1_problem_id.is_(None),
        ).count()
        control.status = "COMPLETE" if unassigned_count == 0 else "READY"
        control.ended = unassigned_count == 0
        event_name = "round1.winners_assigned" if winners else "round1.problem_received_no_bids"
        record_event(db, event_name, actor=current_user, entity_type="problem", entity_id=problem.id, metadata={"winner_team_ids": [winner["team_id"] for winner in winners]})
        db.commit()
        message = (
            f"{len(winners)} actual bidder{'s' if len(winners) != 1 else ''} assigned."
            if winners
            else "No bids received. Problem moved to remaining allocation pool."
        )
        response = {"message": message, "winners": winners, **_round_payload(db, meta)}
        db.close()
    await manager.broadcast_event("round_updated", {
        "round": meta["type"],
        "action": "winners_assigned" if winners else "problem_no_bids",
        "winners": winners,
        "problem": assigned_problem,
    })
    return response


@router.get("/admin/rounds/round-1/assignments")
def get_round_one_assignments(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_admin),
):
    del current_user
    return round1_assignment_management_payload(db)


@router.post("/admin/rounds/round-1/assignments/external-problems/import")
async def import_external_assignment_problems(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_admin),
):
    """Import manual-assignment-only problems using the established Round 1 parser."""
    rows = _read_problem_rows(file.filename or "", await file.read())
    with ROUND1_FINALIZATION_LOCK:
        existing = db.query(ProblemStatement).with_for_update().all()
        existing_by_number: dict[int, list[ProblemStatement]] = {}
        for problem in existing:
            try:
                existing_by_number.setdefault(int(_display_number(problem)), []).append(problem)
            except (TypeError, ValueError):
                continue

        conflicts: list[str] = []
        skipped_duplicates: list[dict] = []
        pending: list[tuple[int, str, str]] = []
        for number, title, description in rows:
            matches = existing_by_number.get(number, [])
            if not matches:
                pending.append((number, title, description))
                continue
            auction_problem = next((problem for problem in matches if problem.round != EXTERNAL_PROBLEM_ROUND), None)
            if auction_problem:
                source = "Round 1" if auction_problem.round == 1 else "Wildcard"
                conflicts.append(
                    f"Problem #{number} already exists in {source}; choose a distinct external problem number."
                )
                continue
            found = matches[0]
            if found.round == EXTERNAL_PROBLEM_ROUND:
                skipped_duplicates.append({
                    "problem_number": str(number),
                    "title": found.title,
                    "reason": f"External problem #{number} already exists; the stored record was not overwritten.",
                })

        if conflicts:
            raise HTTPException(status_code=409, detail=conflicts)

        created: list[ProblemStatement] = []
        try:
            for number, title, description in pending:
                problem = ProblemStatement(
                    ps_number=f"EX-{number}",
                    title=title,
                    description=description,
                    round=EXTERNAL_PROBLEM_ROUND,
                    status="available",
                )
                db.add(problem)
                created.append(problem)
            db.flush()
            record_event(
                db,
                "round1.external_problems_imported",
                actor=current_user,
                metadata={
                    "created": len(created),
                    "skipped_duplicates": len(skipped_duplicates),
                    "filename": file.filename or "",
                },
            )
            db.commit()
        except IntegrityError as exc:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail="An external problem with the same number was imported concurrently. Refresh and retry.",
            ) from exc

    snapshot = round1_assignment_management_payload(db)
    db.close()
    await manager.broadcast_event("external_problems_imported", {
        "created": len(created),
        "skipped_duplicates": len(skipped_duplicates),
    })
    return {
        "message": (
            f"Imported {len(created)} external problem{'s' if len(created) != 1 else ''}. "
            f"Skipped {len(skipped_duplicates)} duplicate{'s' if len(skipped_duplicates) != 1 else ''}."
        ),
        "imported": len(created),
        "skipped_duplicate_count": len(skipped_duplicates),
        "skipped_duplicates": skipped_duplicates,
        **snapshot,
    }


@router.get("/admin/rounds/round-1/assignments/external-problems/sample.csv")
def external_problem_sample(current_user=Depends(get_current_active_admin)):
    del current_user
    sample = (
        "Problem Number,Title,Description\r\n"
        "21,\"Smart Parking Optimization\",\"Design a system that optimizes parking availability and routing.\"\r\n"
        "22,\"Disaster Communication System\",\"Build a resilient communication platform for disaster response.\"\r\n"
    )
    return Response(
        sample,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=external-problems-sample.csv"},
    )


@router.put("/admin/rounds/round-1/assignments/{team_id}")
async def change_round_one_assignment(
    team_id: int,
    payload: ChangeProblemAssignmentRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_admin),
):
    try:
        result = change_round1_problem_assignment(
            db,
            team_id,
            payload.target_problem_id,
            new_balance=payload.new_balance,
            actor=current_user,
        )
    except Round1AssignmentError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except (IntegrityError, OperationalError) as exc:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Another Admin changed Round 1 assignments at the same time. Refresh and retry.",
        ) from exc

    if not result["idempotent"]:
        change = result["change"]
        db.close()
        await manager.broadcast_event("round1_assignment_changed", {
            "team_id": change["team_id"],
            "previous_problem_id": change["previous_problem_id"],
            "problem": change["problem"],
            "coins": change["coins"],
        })
    return {
        "message": "Problem assignment already matches the requested problem."
        if result["idempotent"]
        else (
            f"Problem assigned and balance set to {result['change']['coins']:,} coins."
            if result["change"]["balance_changed"]
            else "Round 1 problem assignment changed. No balance change."
        ),
        **result,
    }


@router.post("/admin/rounds/round-1/problems/{problem_id}/assign")
async def assign_problem_manually(
    problem_id: int,
    payload: ManualProblemAssignmentRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_admin),
):
    control = get_or_create_round_control(db, "ROUND1")
    try:
        result = manually_assign_problem(
            db,
            control,
            problem_id,
            payload.team_ids,
            payload.deduction,
            actor=current_user,
        )
        db.commit()
    except Round1AssignmentError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except (IntegrityError, OperationalError) as exc:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="The assignment changed concurrently. Refresh the Round 1 controls and retry.",
        ) from exc
    if not result["idempotent"]:
        snapshot = event_snapshot(db)
        response = {
            "message": "Manual Round 1 assignment completed.",
            "idempotent": False,
            **_round_payload(db, ROUND_META["round-1"]),
        }
        db.close()
        await manager.broadcast_event("round_updated", {
            "round": "ROUND1",
            "action": "problem_manually_assigned",
            "problem_id": problem_id,
            "assignments": result["assignments"],
        })
        await manager.broadcast_event("event_state_changed", snapshot)
        return response
    return {"message": "Assignment already applied.", "idempotent": True, **_round_payload(db, ROUND_META["round-1"])}


@router.post("/admin/rounds/round-1/problems/{problem_id}/rebid")
async def rebid_problem(
    problem_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_admin),
):
    with ROUND1_FINALIZATION_LOCK:
        unlocked = get_or_create_round_control(db, "ROUND1")
        control = (
            db.query(RoundControl)
            .filter(RoundControl.id == unlocked.id)
            .with_for_update()
            .populate_existing()
            .one()
        )
        problem = (
            db.query(ProblemStatement)
            .filter(ProblemStatement.id == problem_id, ProblemStatement.round == 1)
            .with_for_update()
            .first()
        )
        if not problem:
            raise HTTPException(status_code=404, detail="Round 1 problem not found.")
        if control.current_problem_id == problem.id:
            return _round_payload(db, ROUND_META["round-1"])
        if control.ended:
            raise HTTPException(status_code=409, detail="Round 1 is closed.")
        if control.current_problem_id is not None or control.status in {"PREVIEW", "BIDDING"}:
            raise HTTPException(status_code=409, detail="Complete the current auction before starting a re-bid.")
        capacity = ROUND1_PROBLEM_CAPACITY - db.query(Team).filter(
            Team.round1_problem_id == problem.id
        ).count()
        if capacity <= 0:
            raise HTTPException(status_code=409, detail="This problem already has five assigned teams.")

        # Bid rows have one unique slot per team/problem/round and no attempt id.
        # Reset only this problem's live slate so a prior losing bid cannot win a re-bid.
        db.query(Bid).filter(Bid.ps_id == problem.id, Bid.round == 1).delete(synchronize_session=False)
        transition_event_state(db, "WAITING", validate=False, commit=False)
        problem.status = "current"
        control.current_problem_id = problem.id
        control.status = "READY"
        record_event(
            db,
            "round1.problem_rebid_selected",
            actor=current_user,
            entity_type="problem",
            entity_id=problem.id,
            metadata={"remaining_capacity": capacity},
        )
        db.commit()
        snapshot = event_snapshot(db)
        response = _round_payload(db, ROUND_META["round-1"])
        db.close()
    await manager.broadcast_event("round_updated", {
        "round": "ROUND1",
        "action": "problem_rebid_selected",
        "problem_id": problem_id,
        "remaining_capacity": capacity,
    })
    await manager.broadcast_event("event_state_changed", snapshot)
    return response


@router.post("/admin/rounds/round-1/end")
async def end_round_one(db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    control = get_or_create_round_control(db, "ROUND1")
    if control.ended:
        logger.info("Duplicate Round 1 end request ignored; the round is already closed.")
        return _round_payload(db, ROUND_META["round-1"])
    assigned_count = db.query(Team).filter(Team.round1_problem_id.is_not(None)).count()
    unassigned_count = db.query(Team).filter(
        Team.is_approved.is_(True), Team.is_system_team.is_(False), Team.round1_problem_id.is_(None)
    ).count()
    if control.current_problem_id:
        current = db.query(ProblemStatement).filter(ProblemStatement.id == control.current_problem_id).first()
        if current and current.status == "current":
            current.status = "available"
    control.current_problem_id = None
    control.ended = True
    control.status = "CLOSED"
    transition_event_state(db, "ROUND1_RESULT", validate=False, commit=False)
    record_event(db, "round1.manually_ended", actor=current_user, metadata={
        "assigned_team_count": assigned_count,
        "unassigned_team_count": unassigned_count,
    })
    db.commit()
    snapshot = event_snapshot(db)
    response = _round_payload(db, ROUND_META["round-1"])
    db.close()
    await manager.broadcast_event("round1_ended", {
        "manual": True,
        "assigned_team_count": assigned_count,
        "unassigned_team_count": unassigned_count,
    })
    await manager.broadcast_event("event_state_changed", snapshot)
    return response


@router.post("/admin/rounds/wildcard/applications/open")
async def open_applications(db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    round1 = get_or_create_round_control(db, "ROUND1")
    wildcard = get_or_create_round_control(db, "WILDCARD")
    if not round1.ended:
        raise HTTPException(status_code=409, detail="End Round 1 before opening Wildcard applications.")
    if wildcard.status == "APPLICATIONS_OPEN":
        return _round_payload(db, ROUND_META["wildcard"])
    if wildcard.status != "NOT_STARTED" or wildcard.slot_count is not None:
        raise HTTPException(status_code=409, detail=f"Wildcard applications cannot open from {wildcard.status}.")
    wildcard.applications_open = True
    wildcard.status = "APPLICATIONS_OPEN"
    transition_event_state(db, "WILDCARD_APPLICATION", validate=False, restart=True, commit=False)
    record_event(db, "wildcard.applications_opened", actor=current_user)
    db.commit()
    snapshot = event_snapshot(db)
    response = _round_payload(db, ROUND_META["wildcard"])
    db.close()
    await manager.broadcast_event("event_state_changed", snapshot)
    return response


@router.post("/admin/rounds/wildcard/applications/close")
async def close_applications(db: Session = Depends(get_db), current_user=Depends(get_current_active_admin)):
    control = get_or_create_round_control(db, "WILDCARD")
    sync_expired_event_state(db)
    db.refresh(control)
    if control.status == "APPLICATIONS_CLOSED":
        return _round_payload(db, ROUND_META["wildcard"])
    if control.status != "APPLICATIONS_OPEN":
        raise HTTPException(status_code=409, detail="Wildcard applications are not open.")
    control.applications_open = False
    control.status = "APPLICATIONS_CLOSED"
    game = get_or_create_game_config(db)
    game.auction_timer_end = None
    game.timer_paused = False
    game.timer_paused_remaining_seconds = None
    record_event(db, "wildcard.applications_closed", actor=current_user)
    db.commit()
    response = _round_payload(db, ROUND_META["wildcard"])
    db.close()
    await manager.broadcast_event("wildcard_updated", {"action": "applications_closed"})
    return response


@router.get("/leaderboard/{round_slug}")
def public_round_leaderboard(
    round_slug: str,
    response: FastAPIResponse,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_display),
):
    del current_user
    meta = _meta(round_slug)
    response.headers["Cache-Control"] = "no-store"
    event_config = get_or_create_event_config(db)
    if meta["type"] == "WILDCARD":
        control = get_or_create_round_control(db, "WILDCARD")
        return {
            "round": "WILDCARD",
            "label": "Wildcard Slot Auction",
            "slot_count": control.slot_count,
            "finalized": control.status in {"PROBLEM_SELECTION", "COMPLETE"},
            "active": control.status == "BIDDING_OPEN",
            "base_price": event_config.wildcard_starting_bid,
            "rows": ranking_payload(db, control),
        }
    control = get_or_create_round_control(db, "ROUND1")
    query = db.query(Bid).filter(Bid.round == 1)
    if control.current_problem_id:
        query = query.filter(Bid.ps_id == control.current_problem_id)
    bids = query.order_by(Bid.amount.desc(), Bid.timestamp.asc(), Bid.team_id.asc()).all()
    highest_by_team = {}
    for bid in bids:
        if bid.team_id not in highest_by_team or bid.amount > highest_by_team[bid.team_id].amount:
            highest_by_team[bid.team_id] = bid
    rows = []
    for team_id, bid in highest_by_team.items():
        team = db.query(Team).filter(Team.id == team_id).first()
        if team:
            rows.append({"team_id": team.id, "team_name": team.team_name, "value": bid.amount, "problem_id": bid.ps_id})
    rows.sort(key=lambda row: (-row["value"], row["team_id"]))
    return {
        "round": meta["type"],
        "label": meta["label"],
        "active": control.status == "BIDDING",
        "finalized": control.ended or control.status == "CLOSED",
        "base_price": event_config.round1_minimum_bid,
        "rows": [{**row, "rank": index} for index, row in enumerate(rows, start=1)],
    }
